#!/usr/bin/env python3
from __future__ import annotations

import argparse
import random
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = "4.14作文前筛.xlsx"
DEFAULT_OUTPUT = "outputs/作文随机抽样结果.xlsx"
DEFAULT_SEED = 20260414

SHEET_SPECS = [
    ("J1", "J1记对我影响最大的一个人"),
    ("J2", "J2我的一个假期"),
    ("Y1", "Y1吸烟对个人健康和公众利益的影响"),
    ("Y2", "Y2父母是孩子的第一任老师"),
]

SCORE_SEGMENTS = [55, 60, 65, 70, 75, 80, 85, 90, 95]
SOURCE_HEADERS = ["作文编码", "国籍", "作文题目", "作文分数", "体裁"]
OUTPUT_HEADERS = ["篇名代码", "篇名", "作文编码", "国籍", "作文题目", "作文分数", "体裁"]
EXCLUDED_REGION_COUNTRIES = {
    "香港",
    "香港地区",
    "中国香港",
    "澳门",
    "澳门地区",
    "中国澳门",
    "台湾",
    "台湾地区",
    "中国台湾",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按分数段在 J1/J2/Y1/Y2 四个作文 sheet 中均衡随机抽样。",
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help=f"输入 Excel 文件，默认：{DEFAULT_INPUT}")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"输出 Excel 文件，默认：{DEFAULT_OUTPUT}")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help=f"随机种子，默认：{DEFAULT_SEED}")
    return parser.parse_args()


def normalize_score(value: Any, sheet_name: str, row_number: int) -> int:
    if value is None or value == "":
        raise ValueError(f"{sheet_name}!D{row_number} 缺少作文分数")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{sheet_name}!D{row_number} 作文分数不是整数：{value!r}") from exc


def essay_title_from_sheet(sheet_name: str, code: str) -> str:
    if not sheet_name.startswith(code):
        raise ValueError(f"sheet 名 `{sheet_name}` 未以篇名代码 `{code}` 开头")
    title = sheet_name[len(code) :]
    if not title:
        raise ValueError(f"sheet 名 `{sheet_name}` 去掉代码后没有篇名")
    return title


def normalize_country(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def excluded_country_reason(value: Any) -> str | None:
    country = normalize_country(value)
    if not country:
        return "缺少国籍"
    if country in EXCLUDED_REGION_COUNTRIES:
        return "港澳台地区"
    if "少数民族" in country or country.endswith("族"):
        return "中国少数民族"
    return None


def read_records(
    input_path: Path,
) -> tuple[dict[str, dict[int, list[dict[str, Any]]]], Counter[str], Counter[tuple[str, int]]]:
    if not input_path.exists():
        raise FileNotFoundError(f"找不到输入文件：{input_path}")

    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        records_by_code_score: dict[str, dict[int, list[dict[str, Any]]]] = {
            code: {score: [] for score in SCORE_SEGMENTS} for code, _ in SHEET_SPECS
        }
        excluded_by_reason: Counter[str] = Counter()
        excluded_by_code_score: Counter[tuple[str, int]] = Counter()

        missing_sheets = [sheet_name for _, sheet_name in SHEET_SPECS if sheet_name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError("输入文件缺少必要 sheet：" + "、".join(missing_sheets))

        for code, sheet_name in SHEET_SPECS:
            sheet = workbook[sheet_name]
            header = [sheet.cell(row=1, column=col).value for col in range(1, 6)]
            if header != SOURCE_HEADERS:
                raise ValueError(f"{sheet_name} 第 1 行表头应为 {SOURCE_HEADERS}，实际为 {header}")

            essay_title = essay_title_from_sheet(sheet_name, code)
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                essay_id, country, topic, score_value, genre = row[:5]
                if all(value in (None, "") for value in (essay_id, country, topic, score_value, genre)):
                    continue

                score = normalize_score(score_value, sheet_name, row_number)
                if score not in SCORE_SEGMENTS:
                    raise ValueError(
                        f"{sheet_name}!D{row_number} 出现未配置分数段 {score}，"
                        f"当前脚本只处理：{SCORE_SEGMENTS}"
                    )

                exclude_reason = excluded_country_reason(country)
                if exclude_reason is not None:
                    excluded_by_reason[exclude_reason] += 1
                    excluded_by_code_score[(code, score)] += 1
                    continue

                records_by_code_score[code][score].append(
                    {
                        "篇名代码": code,
                        "篇名": essay_title,
                        "作文编码": essay_id,
                        "国籍": normalize_country(country),
                        "作文题目": topic,
                        "作文分数": score,
                        "体裁": genre,
                    }
                )

        return records_by_code_score, excluded_by_reason, excluded_by_code_score
    finally:
        workbook.close()


def balanced_sample_by_country(
    candidates: list[dict[str, Any]],
    sample_size: int,
    rng: random.Random,
    global_country_counts: Counter[str],
) -> list[dict[str, Any]]:
    if sample_size == 0:
        return []
    if sample_size > len(candidates):
        raise ValueError(f"候选数 {len(candidates)} 小于抽样数 {sample_size}")

    by_country: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in candidates:
        by_country[record["国籍"]].append(record)

    for country_records in by_country.values():
        rng.shuffle(country_records)

    countries = sorted(by_country)
    if len(countries) > sample_size:
        tie_breakers = {country: rng.random() for country in countries}
        countries = sorted(
            countries,
            key=lambda country: (global_country_counts[country], tie_breakers[country], country),
        )[:sample_size]

    allocation = {country: 1 for country in countries}
    remaining = sample_size - len(countries)
    while remaining > 0:
        eligible = [country for country in countries if allocation[country] < len(by_country[country])]
        if not eligible:
            raise ValueError("国籍均衡抽样时没有可用候选，可能是候选数据不完整")

        tie_breakers = {country: rng.random() for country in eligible}
        country = min(
            eligible,
            key=lambda item: (
                global_country_counts[item] + allocation[item],
                allocation[item],
                global_country_counts[item],
                tie_breakers[item],
                item,
            ),
        )
        allocation[country] += 1
        remaining -= 1

    selected: list[dict[str, Any]] = []
    for country in countries:
        selected.extend(by_country[country][: allocation[country]])
    rng.shuffle(selected)
    return selected


def sample_records(
    records_by_code_score: dict[str, dict[int, list[dict[str, Any]]]],
    seed: int,
) -> tuple[list[dict[str, Any]], dict[int, int], dict[int, dict[str, int]]]:
    rng = random.Random(seed)
    selected: list[dict[str, Any]] = []
    x_by_score: dict[int, int] = {}
    counts_by_score: dict[int, dict[str, int]] = {}
    selected_by_group: dict[tuple[int, str], list[dict[str, Any]]] = {}
    global_country_counts: Counter[str] = Counter()

    for score in SCORE_SEGMENTS:
        counts = {code: len(records_by_code_score[code][score]) for code, _ in SHEET_SPECS}
        x_value = min(counts.values())
        counts_by_score[score] = counts
        x_by_score[score] = x_value

    group_keys = [(score, code) for score in SCORE_SEGMENTS for code, _ in SHEET_SPECS]
    rng.shuffle(group_keys)
    for score, code in group_keys:
        candidates = records_by_code_score[code][score]
        group_selected = balanced_sample_by_country(candidates, x_by_score[score], rng, global_country_counts)
        selected_by_group[(score, code)] = group_selected
        global_country_counts.update(row["国籍"] for row in group_selected)

    for score in SCORE_SEGMENTS:
        for code, _ in SHEET_SPECS:
            selected.extend(selected_by_group[(score, code)])

    return selected, x_by_score, counts_by_score


def validate_selection(
    selected: list[dict[str, Any]],
    x_by_score: dict[int, int],
    records_by_code_score: dict[str, dict[int, list[dict[str, Any]]]],
) -> None:
    if OUTPUT_HEADERS != ["篇名代码", "篇名", "作文编码", "国籍", "作文题目", "作文分数", "体裁"]:
        raise AssertionError("输出表头配置被意外修改")

    expected_total = sum(x_by_score[score] * len(SHEET_SPECS) for score in SCORE_SEGMENTS)
    if len(selected) != expected_total:
        raise AssertionError(f"输出行数应为 {expected_total}，实际为 {len(selected)}")

    ids = [row["作文编码"] for row in selected]
    duplicate_ids = [essay_id for essay_id, count in Counter(ids).items() if count > 1]
    if duplicate_ids:
        raise AssertionError("抽样结果作文编码重复：" + "、".join(map(str, duplicate_ids[:10])))

    actual = Counter((row["作文分数"], row["篇名代码"]) for row in selected)
    selected_by_code_score: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in selected:
        if excluded_country_reason(row["国籍"]) is not None:
            raise AssertionError(f"抽样结果包含应排除国籍：{row['国籍']!r}")
        selected_by_code_score[(row["篇名代码"], row["作文分数"])].append(row)

    for score in SCORE_SEGMENTS:
        for code, _ in SHEET_SPECS:
            count = actual[(score, code)]
            if count != x_by_score[score]:
                raise AssertionError(f"{score} 分 {code} 应抽 {x_by_score[score]} 条，实际 {count} 条")

            candidate_country_count = len({row["国籍"] for row in records_by_code_score[code][score]})
            selected_country_count = len({row["国籍"] for row in selected_by_code_score[(code, score)]})
            max_possible_country_count = min(x_by_score[score], candidate_country_count)
            if selected_country_count != max_possible_country_count:
                raise AssertionError(
                    f"{score} 分 {code} 应覆盖 {max_possible_country_count} 个国籍，"
                    f"实际覆盖 {selected_country_count} 个"
                )


def write_output(output_path: Path, selected: list[dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "抽样结果"
    sheet.append(OUTPUT_HEADERS)

    for row in selected:
        values = [row[header] for header in OUTPUT_HEADERS]
        values[2] = str(values[2])
        sheet.append(values)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 10,
        "B": 36,
        "C": 28,
        "D": 14,
        "E": 42,
        "F": 10,
        "G": 10,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        row[2].number_format = "@"
        row[2].quotePrefix = True
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_HEADERS))}{sheet.max_row}"

    workbook.save(output_path)


def print_summary(
    counts_by_score: dict[int, dict[str, int]],
    x_by_score: dict[int, int],
    total_selected: int,
    output_path: Path,
    seed: int,
    selected: list[dict[str, Any]],
    excluded_by_reason: Counter[str],
    excluded_by_code_score: Counter[tuple[str, int]],
) -> None:
    code_order = [code for code, _ in SHEET_SPECS]
    print(f"随机种子：{seed}")
    if excluded_by_reason:
        print("已排除候选：")
        for reason, count in excluded_by_reason.most_common():
            print(f"- {reason}：{count} 条")
        for score in SCORE_SEGMENTS:
            details = [
                f"{code}={excluded_by_code_score[(code, score)]}"
                for code in code_order
                if excluded_by_code_score[(code, score)] > 0
            ]
            if details:
                print(f"  {score}分：" + "，".join(details))
    print("分数段抽样汇总：")
    for score in SCORE_SEGMENTS:
        counts_text = "，".join(f"{code}={counts_by_score[score][code]}" for code in code_order)
        actual_selected = x_by_score[score] * len(SHEET_SPECS)
        print(f"- {score}分：{counts_text}；X={x_by_score[score]}；实际抽样={actual_selected}")
    print("抽样结果国籍分布 Top 15：")
    for country, count in Counter(row["国籍"] for row in selected).most_common(15):
        print(f"- {country}：{count}")
    print(f"总抽样行数：{total_selected}")
    print(f"输出文件：{output_path}")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    if input_path.resolve() == output_path.resolve():
        raise ValueError("输出路径不能和输入文件相同，避免覆盖原始 workbook")

    records_by_code_score, excluded_by_reason, excluded_by_code_score = read_records(input_path)
    selected, x_by_score, counts_by_score = sample_records(records_by_code_score, args.seed)
    validate_selection(selected, x_by_score, records_by_code_score)
    write_output(output_path, selected)
    print_summary(
        counts_by_score,
        x_by_score,
        len(selected),
        output_path,
        args.seed,
        selected,
        excluded_by_reason,
        excluded_by_code_score,
    )


if __name__ == "__main__":
    main()
