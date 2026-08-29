#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import random
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = "outputs/作文随机抽样结果_去掉75分.xlsx"
DEFAULT_OUTPUT_DIR = "ori_text"
BASE_URL = "https://hsk.blcu.edu.cn/api/v1/resource/text/"
EXPECTED_TOTAL = 620
ALLOWED_CODES = {"J1", "J2", "Y1", "Y2"}
RETRY_HTTP_STATUS = {403, 429, 500, 502, 503, 504}
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


@dataclass(frozen=True)
class EssayRecord:
    index: int
    code: str
    essay_id: str


@dataclass
class FetchFailure:
    index: int
    code: str
    essay_id: str
    error: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="批量抓取 HSK 作文标注文本，保存到 ./ori_text/{代码}/{作文编码}.txt。",
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help=f"输入 Excel，默认：{DEFAULT_INPUT}")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help=f"输出目录，默认：{DEFAULT_OUTPUT_DIR}")
    parser.add_argument("--min-delay", type=float, default=2.0, help="每次请求后的最小随机等待秒数，默认：2.0")
    parser.add_argument("--max-delay", type=float, default=5.0, help="每次请求后的最大随机等待秒数，默认：5.0")
    parser.add_argument("--max-retries", type=int, default=5, help="单篇失败最大重试次数，默认：5")
    parser.add_argument("--timeout", type=float, default=30.0, help="单次请求超时秒数，默认：30.0")
    parser.add_argument("--limit", type=int, default=None, help="只处理前 N 条，用于小样本验证")
    parser.add_argument("--force", action="store_true", help="覆盖已存在的 txt；默认跳过已有文件")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.min_delay < 0 or args.max_delay < 0:
        raise ValueError("--min-delay 和 --max-delay 不能为负数")
    if args.min_delay > args.max_delay:
        raise ValueError("--min-delay 不能大于 --max-delay")
    if args.max_retries < 0:
        raise ValueError("--max-retries 不能为负数")
    if args.timeout <= 0:
        raise ValueError("--timeout 必须大于 0")
    if args.limit is not None and args.limit <= 0:
        raise ValueError("--limit 必须大于 0")


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_records(input_path: Path, limit: int | None) -> list[EssayRecord]:
    if not input_path.exists():
        raise FileNotFoundError(f"找不到输入 Excel：{input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if "抽样结果" not in workbook.sheetnames:
            raise ValueError(f"{input_path} 缺少 `抽样结果` sheet")

        sheet = workbook["抽样结果"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        try:
            code_col = headers.index("篇名代码")
            essay_id_col = headers.index("作文编码")
        except ValueError as exc:
            raise ValueError(f"源表缺少必要列：篇名代码、作文编码；实际表头：{headers}") from exc

        records: list[EssayRecord] = []
        seen_ids: set[str] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value in (None, "") for value in row):
                continue

            code = normalize_cell(row[code_col])
            essay_id = normalize_cell(row[essay_id_col])
            if code not in ALLOWED_CODES:
                raise ValueError(f"{sheet.title}!A{row_number} 篇名代码不在 {sorted(ALLOWED_CODES)} 中：{code!r}")
            if not essay_id:
                raise ValueError(f"{sheet.title}!C{row_number} 缺少作文编码")
            if essay_id in seen_ids:
                raise ValueError(f"作文编码重复：{essay_id}")
            seen_ids.add(essay_id)
            records.append(EssayRecord(index=len(records) + 1, code=code, essay_id=essay_id))

        if len(records) != EXPECTED_TOTAL:
            raise ValueError(f"源表应有 {EXPECTED_TOTAL} 条作文，实际读取 {len(records)} 条")

        return records[:limit] if limit is not None else records
    finally:
        workbook.close()


def output_path_for(output_dir: Path, record: EssayRecord) -> Path:
    return output_dir / record.code / f"{record.essay_id}.txt"


def fetch_text(essay_id: str, timeout: float, max_retries: int) -> str:
    url = BASE_URL + urllib.parse.quote(essay_id)
    last_error = ""

    for attempt in range(max_retries + 1):
        request = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                status = response.getcode()
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            status = exc.code
            detail = exc.read().decode("utf-8", errors="replace")
            last_error = f"HTTP {status}: {detail[:300]}"
            if status not in RETRY_HTTP_STATUS or attempt == max_retries:
                raise RuntimeError(last_error) from exc
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt == max_retries:
                raise RuntimeError(last_error) from exc
        else:
            if status != 200:
                last_error = f"HTTP {status}: {body[:300]}"
                if status not in RETRY_HTTP_STATUS or attempt == max_retries:
                    raise RuntimeError(last_error)
            else:
                try:
                    payload = json.loads(body)
                except json.JSONDecodeError as exc:
                    raise RuntimeError(f"JSON 解析失败：{body[:300]}") from exc

                if payload.get("code") != 0:
                    raise RuntimeError(f"接口返回失败：{payload}")
                data = payload.get("data")
                if not isinstance(data, str) or data == "":
                    raise RuntimeError(f"接口未返回有效文本：{payload}")
                return data

        backoff = (10 * (2**attempt)) + random.uniform(0.5, 3.0)
        print(f"  请求失败，{backoff:.1f}s 后重试 {attempt + 1}/{max_retries}：{last_error}", flush=True)
        time.sleep(backoff)

    raise RuntimeError(last_error or "未知请求失败")


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as tmp:
        tmp.write(text)
        tmp_path = Path(tmp.name)
    tmp_path.replace(path)


def sleep_between_requests(args: argparse.Namespace) -> None:
    if args.max_delay == 0:
        return
    delay = random.uniform(args.min_delay, args.max_delay)
    print(f"  等待 {delay:.1f}s，降低触发反爬风险", flush=True)
    time.sleep(delay)


def write_failures(output_dir: Path, failures: list[FetchFailure]) -> Path | None:
    failure_path = output_dir / "fetch_failures.csv"
    if not failures:
        if failure_path.exists():
            failure_path.unlink()
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    with failure_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["index", "code", "essay_id", "error"])
        writer.writeheader()
        for failure in failures:
            writer.writerow(
                {
                    "index": failure.index,
                    "code": failure.code,
                    "essay_id": failure.essay_id,
                    "error": failure.error,
                }
            )
    return failure_path


def main() -> int:
    args = parse_args()
    validate_args(args)

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    records = read_records(input_path, args.limit)
    total = len(records)

    print(f"输入文件：{input_path}")
    print(f"输出目录：{output_dir}")
    print(f"本次处理：{total} 条")
    print(f"请求间隔：{args.min_delay:.1f}-{args.max_delay:.1f}s；最大重试：{args.max_retries}")

    saved = 0
    skipped = 0
    failures: list[FetchFailure] = []

    for position, record in enumerate(records, start=1):
        path = output_path_for(output_dir, record)
        progress = f"[{position}/{total}] {record.code} {record.essay_id}"

        if path.exists() and not args.force:
            skipped += 1
            print(f"{progress} 已存在，跳过：{path}", flush=True)
            continue

        print(f"{progress} 开始抓取", flush=True)
        try:
            text = fetch_text(record.essay_id, args.timeout, args.max_retries)
            atomic_write_text(path, text)
        except Exception as exc:  # noqa: BLE001 - keep batch running and write CSV for review.
            failures.append(FetchFailure(record.index, record.code, record.essay_id, str(exc)))
            print(f"{progress} 失败：{exc}", flush=True)
        else:
            saved += 1
            print(f"{progress} 已保存：{path}", flush=True)
        finally:
            if position < total:
                sleep_between_requests(args)

    failure_path = write_failures(output_dir, failures)
    print(f"完成：保存 {saved}，跳过 {skipped}，失败 {len(failures)}")
    if failure_path is not None:
        print(f"失败清单：{failure_path}")
        return 1
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\n用户中断，可稍后重新运行脚本继续抓取已缺失文件。", file=sys.stderr)
        raise SystemExit(130)
