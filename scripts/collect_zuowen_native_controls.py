#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import random
import re
import statistics
import sys
import time
from collections import Counter
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Iterable, Sequence
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen


SEARCH_ENDPOINT = "https://www.zuowen.com/search/api/multiple_search/zhanqun/search"
DEFAULT_LEARNER_WORKBOOK = "作文词性统计宽表.xlsx"
DEFAULT_CACHE_DIR = "native_cache"
DEFAULT_ORI_DIR = "native_ori_text"
DEFAULT_CLEAN_DIR = "native_clean_text"
DEFAULT_OUTPUT_DIR = "outputs/native_control"
DEFAULT_MANUAL_REVIEW = "resources/native_control_manual_review.csv"
DEFAULT_TARGET = 50
DEFAULT_SEED = 20260830
HAN_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
DATE_RE = re.compile(r"(20\d{2}|19\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})")
NOMINAL_LENGTH_RE = re.compile(r"[_（(](\d{2,4})\s*字?[）)]?|_(\d{2,4})字")
ARTICLE_ID_RE = re.compile(r"/e/\d{8}/([0-9a-fA-F]+)\.shtml")
ART_ID_RE = re.compile(r"\bartID\s*=\s*['\"]([^'\"]+)['\"]")
SERIES_SUFFIX_RE = re.compile(r"(?:\s|_)*(?:第?[一二三四五六七八九十百\d]+篇?|[一二三四五六七八九十百\d]+)$")

CATEGORY_GROUP_URLS: tuple[tuple[tuple[str, ...], str], ...] = (
    (("NJ1",), "https://www.zuowen.com/gaozhong/gaoyi/xieren/"),
    (("NJ1",), "https://www.zuowen.com/gaozhong/gaoer/xieren/"),
    (("NJ1",), "https://www.zuowen.com/gaozhong/gaosan/xieren/"),
    (("NJ2",), "https://www.zuowen.com/gaozhong/gaoyi/xunshi/"),
    (("NJ2",), "https://www.zuowen.com/gaozhong/gaoer/xushi/"),
    (("NJ2",), "https://www.zuowen.com/gaozhong/gaosan/xushi/"),
    (("NY1", "NY2"), "https://www.zuowen.com/gaozhong/gaoyi/yilunwen/"),
    (("NY1", "NY2"), "https://www.zuowen.com/gaozhong/gaoer/yilunwen/"),
    (("NY1", "NY2"), "https://www.zuowen.com/gaozhong/gaosan/yilunwen/"),
)

LIBRARY_GROUP_GENRES: tuple[tuple[tuple[str, ...], int, str], ...] = (
    (("NJ1",), 1, "写人作文"),
    (("NJ2",), 2, "叙事作文"),
    (("NY1", "NY2"), 9, "议论文"),
    (("NY1", "NY2"), 6, "话题作文"),
)
LIBRARY_LENGTH_CODES = tuple(range(7, 19))  # 200至750字

MIDDLE_SCHOOL_NY2_QUERIES = (
    "初一 家庭教育 议论文",
    "初二 家庭教育 议论文",
    "初三 家庭教育 议论文",
    "初中 父母教育 议论文",
    "初中 父母 教导",
    "初中 父母 放手",
    "初中 亲子关系",
    "初中 父母 溺爱",
    "初中 家风 议论文",
    "初中 言传身教",
)
MIDDLE_SCHOOL_NY2_CATEGORY_URLS = (
    "https://www.zuowen.com/chuzhong/chuyi/yilunwen/",
    "https://www.zuowen.com/chuzhong/chuer/yilunwen/",
    "https://www.zuowen.com/chuzhong/chusan/yilunwen/",
)

EXCLUDED_TERMS = (
    "满分作文",
    "范文",
    "作文题目",
    "作文素材",
    "写作素材",
    "写作指导",
    "作文指导",
    "点评",
    "解析",
    "读后感",
    "演讲稿",
    "小说",
    "诗歌",
    "英语作文",
    "说明文",
    "预测",
    "零分作文",
)

BOILERPLATE_PATTERNS = (
    re.compile(r"作文网(?:专稿|原创)?\s*未经允许不得转载"),
    re.compile(r"未经允许不得转载"),
    re.compile(r"^(?:E度网?专稿|E度)$"),
    re.compile(r"^作文网专稿[,，]?$"),
    re.compile(r"^本文系.*(?:用户原创文章|原创文章).*?(?:转载.*)?$"),
    re.compile(r"^(?:中小学生|中学生)写作指导、写作素材、优秀作文以及有奖活动.*$"),
    re.compile(r"^尽在[“\"]?作文网[”\"]?微信公众号.*$"),
    re.compile(r"^指导老师[:：].*$"),
    re.compile(r"扫描二维码.*"),
    re.compile(r"关注作文网.*"),
    re.compile(r"ID[:：]?www_zuowen_com", re.I),
    re.compile(r"(?:责任编辑|编辑)[:：].*"),
    re.compile(r"^(?:高一|高二|高三|高中).*[:：]\s*$"),
)


@dataclass(frozen=True)
class TopicRule:
    native_code: str
    learner_code: str
    learner_title: str
    genre: str
    queries: tuple[str, ...]
    exact_terms: tuple[str, ...]
    near_terms: tuple[str, ...]
    extended_terms: tuple[str, ...]
    context_terms: tuple[str, ...] = ()
    broad_terms: tuple[str, ...] = ()


TOPIC_RULES: dict[str, TopicRule] = {
    "NJ1": TopicRule(
        native_code="NJ1",
        learner_code="J1",
        learner_title="记对我影响最大的一个人",
        genre="记叙文",
        queries=(
            "高一 对我影响最大的人",
            "高二 对我影响最大的人",
            "高三 对我影响最大的人",
            "高中 改变我的人",
            "高中 影响我的老师",
            "高中 影响我的父亲",
            "高中 我的榜样",
            "高一 写人 父亲",
            "高二 写人 老师",
            "对我影响最大的人",
            "改变我的人",
            "影响我的老师",
            "我的榜样",
        ),
        exact_terms=("对我影响最大", "影响我的人", "改变了我", "影响了我", "使我改变"),
        near_terms=("榜样", "引领我", "教会了我", "让我懂得", "给我的影响", "改变我"),
        extended_terms=("父亲", "爸爸", "母亲", "妈妈", "老师", "朋友", "同学", "祖父", "祖母", "爷爷", "奶奶"),
        context_terms=("我", "记得", "后来", "从此", "那一次", "成长"),
    ),
    "NJ2": TopicRule(
        native_code="NJ2",
        learner_code="J2",
        learner_title="我的一个假期",
        genre="记叙文",
        queries=(
            "高一 假期经历",
            "高二 假期经历",
            "高三 假期经历",
            "高中 难忘的假期",
            "高一 寒假生活",
            "高二 暑假生活",
            "高三 假期旅行",
            "高中 旅行经历",
            "高中 旅游记",
            "假期经历",
            "难忘的假期",
            "寒假生活",
            "暑假生活",
            "旅行经历",
        ),
        exact_terms=("假期经历", "难忘的假期", "我的假期", "假期生活", "一次假期"),
        near_terms=("寒假", "暑假", "放假", "假日", "假期"),
        extended_terms=("旅行", "旅游", "游记", "春游", "秋游", "春节", "夏令营"),
        context_terms=("经历", "难忘", "那天", "我们", "来到", "回家", "出发"),
        broad_terms=(
            "亲身经历",
            "难忘",
            "经历",
            "回忆",
            "记忆",
            "第一次",
            "一件事",
            "故事",
            "成长",
            "实践",
            "活动",
            "相遇",
            "告别",
            "旅途",
        ),
    ),
    "NY1": TopicRule(
        native_code="NY1",
        learner_code="Y1",
        learner_title="吸烟对个人健康和公众利益的影响",
        genre="议论文",
        queries=(
            "高一 吸烟危害 议论文",
            "高二 吸烟危害 议论文",
            "高三 吸烟危害 议论文",
            "高中 禁烟 议论文",
            "高中 二手烟",
            "高中 健康生活 议论文",
            "高一 健康 议论文",
            "高二 公共卫生 议论文",
            "高三 健康行为 议论文",
            "高中 珍爱生命 健康",
            "高中 不良习惯 危害 议论文",
            "高中 公共场所 文明 议论文",
            "高中 环境污染 健康 议论文",
            "高中 个人行为 社会影响 议论文",
            "高中 社会公德 公共利益 议论文",
            "高中 健康生活方式 议论文",
            "吸烟危害",
            "青少年吸烟",
            "禁止吸烟",
            "健康生活",
            "公共卫生",
        ),
        exact_terms=("吸烟", "抽烟", "禁烟", "戒烟", "二手烟", "烟草", "香烟"),
        near_terms=("健康生活", "危害健康", "珍爱生命", "健康习惯", "不良习惯"),
        extended_terms=(
            "公共卫生",
            "公众健康",
            "食品安全",
            "保护健康",
            "健康中国",
            "文明行为",
            "公共利益",
            "生命健康",
            "环境污染",
            "空气污染",
            "公共场所",
            "社会公德",
            "文明素养",
            "不文明行为",
            "个人行为",
            "生活方式",
        ),
        context_terms=(
            "危害",
            "影响",
            "应该",
            "反对",
            "禁止",
            "个人",
            "社会",
            "公众",
            "健康",
            "责任",
            "他人",
            "环境",
        ),
        broad_terms=(
            "社会责任",
            "个人责任",
            "公共责任",
            "责任",
            "社会公德",
            "道德",
            "文明",
            "规则",
            "自律",
            "环保",
            "保护环境",
            "绿色生活",
            "低碳",
            "节约",
            "生态文明",
            "交通安全",
            "安全",
            "拒绝毒品",
            "远离毒品",
            "尊重生命",
            "健康成长",
            "心理健康",
            "网瘾",
            "关心他人",
            "尊重他人",
            "诚信",
            "良知",
            "公正",
            "公平",
            "廉洁",
            "节俭",
            "劳动",
            "手机",
            "网络",
            "诱惑",
            "素质",
            "高尚",
            "宽容",
            "尊重",
            "合作",
            "秩序",
            "公益",
            "公民",
        ),
    ),
    "NY2": TopicRule(
        native_code="NY2",
        learner_code="Y2",
        learner_title="父母是孩子的第一任老师",
        genre="议论文",
        queries=(
            "高一 父母是孩子的第一任老师",
            "高二 父母是孩子的第一任老师",
            "高三 父母是孩子的第一任老师",
            "高中 家庭教育 议论文",
            "高中 言传身教 议论文",
            "高中 父母榜样",
            "高中 父母教育孩子",
            "高一 家庭教育",
            "高二 家庭教育",
            "高三 家庭教育",
            "父母是孩子的第一任老师",
            "家庭教育",
            "言传身教",
            "父母教育",
            "教育孩子",
            "父母榜样",
            "以身作则",
            "高中 父母 放手",
            "高中 父母 教导",
            "高中 父母 溺爱",
            "高中 亲子关系",
            "高中 家风",
            "高中 父母 陪伴 成长",
            "高中 家庭影响 成长 议论文",
            "高中 家庭环境 成长 议论文",
            "高中 父母与子女 议论文",
            "高中 亲子沟通 议论文",
            "高中 父爱 母爱 成长 议论文",
            "高中 感恩父母 议论文",
            "高中 理解父母 议论文",
        ),
        exact_terms=(
            "第一任老师",
            "父母是孩子",
            "父母的教育",
            "父母教育",
            "教育孩子",
            "父母教会",
            "父母的教导",
            "家长的教育",
        ),
        near_terms=(
            "家庭教育",
            "言传身教",
            "父母榜样",
            "家长教育",
            "父母影响",
            "亲子关系",
            "父母放手",
            "父母溺爱",
            "父母陪伴",
            "家风",
            "家庭影响",
            "家庭环境",
            "亲子沟通",
            "父母与孩子",
            "父母与子女",
        ),
        extended_terms=(
            "父母",
            "家长",
            "家庭",
            "孩子成长",
            "子女",
            "榜样作用",
            "以身作则",
            "爸妈",
            "爸爸妈妈",
            "管教",
            "溺爱",
            "父爱",
            "母爱",
            "亲情",
            "成长环境",
            "陪伴成长",
            "感恩父母",
            "理解父母",
            "代沟",
        ),
        context_terms=(
            "教育",
            "成长",
            "影响",
            "榜样",
            "责任",
            "应该",
            "孩子",
            "独立",
            "放手",
            "教导",
            "身教",
            "自由",
            "束缚",
            "沟通",
            "关爱",
            "陪伴",
            "家庭环境",
        ),
        broad_terms=(
            "教育",
            "成长",
            "独立",
            "青春",
            "亲情",
            "感恩",
            "陪伴",
            "理解",
            "尊重",
            "自由",
            "关爱",
            "学校教育",
            "素质教育",
            "惩戒",
            "教师",
            "老师",
            "学生",
            "少年",
            "青年",
            "逆境",
            "挫折",
            "成才",
            "成人",
            "学习",
            "学校",
        ),
    ),
}


@dataclass
class SearchCandidate:
    native_code: str
    url: str
    title: str
    search_date: str
    query: str
    source_label: str
    nominal_length: int | None
    pre_tier: str
    pre_score: float
    verified_high_school_hint: bool = False
    library_verified: bool = False
    verified_grade: str = ""
    verified_genre: str = ""
    middle_school_backup: bool = False


@dataclass
class ParsedArticle:
    url: str
    article_id: str
    title: str
    breadcrumb: str
    grade: str
    genre_label: str
    published_date: str
    source_label: str
    raw_text: str
    clean_text: str
    han_count: int
    text_hash: str


@dataclass
class CandidateAudit:
    native_code: str
    learner_code: str
    url: str
    article_id: str = ""
    title: str = ""
    published_date: str = ""
    grade: str = ""
    school_stage: str = ""
    genre_label: str = ""
    source_label: str = ""
    han_count: int = 0
    topic_tier: str = ""
    topic_score: float = 0.0
    status: str = ""
    reason: str = ""
    text_hash: str = ""
    query: str = ""
    nominal_length: int | None = None
    article: ParsedArticle | None = field(default=None, repr=False, compare=False)


class ZuowenHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.stack: list[tuple[str, set[str]]] = []
        self.capture_path_depth: int | None = None
        self.capture_body_depth: int | None = None
        self.capture_title_depth: int | None = None
        self.capture_meta_depth: int | None = None
        self.path_parts: list[str] = []
        self.body_parts: list[str] = []
        self.title_parts: list[str] = []
        self.meta_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr_map = {key: value or "" for key, value in attrs}
        classes = set(attr_map.get("class", "").split())
        self.stack.append((tag, classes))
        depth = len(self.stack)
        if "path" in classes and self.capture_path_depth is None:
            self.capture_path_depth = depth
        if "con_content" in classes and self.capture_body_depth is None:
            self.capture_body_depth = depth
        if tag == "h1" and "h_title" in classes and self.capture_title_depth is None:
            self.capture_title_depth = depth
        if tag == "p" and "text-align:center" in attr_map.get("style", "").replace(" ", ""):
            self.capture_meta_depth = depth
        if self.capture_body_depth is not None and tag in {"br", "p", "div"}:
            self.body_parts.append("\n")

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if self.capture_body_depth is not None and tag == "br":
            self.body_parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        depth = len(self.stack)
        if self.capture_body_depth is not None and tag in {"p", "div"}:
            self.body_parts.append("\n")
        if self.capture_path_depth == depth:
            self.capture_path_depth = None
        if self.capture_body_depth == depth:
            self.capture_body_depth = None
        if self.capture_title_depth == depth:
            self.capture_title_depth = None
        if self.capture_meta_depth == depth:
            self.capture_meta_depth = None
        if self.stack:
            self.stack.pop()

    def handle_data(self, data: str) -> None:
        if self.capture_path_depth is not None:
            self.path_parts.append(data)
        if self.capture_body_depth is not None:
            self.body_parts.append(data)
        if self.capture_title_depth is not None:
            self.title_parts.append(data)
        if self.capture_meta_depth is not None:
            self.meta_parts.append(data)


class SlowHttpClient:
    def __init__(
        self,
        cache_dir: Path,
        min_delay: float,
        max_delay: float,
        max_retries: int,
        seed: int,
        timeout: float = 25.0,
    ) -> None:
        self.cache_dir = cache_dir
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.min_delay = min_delay
        self.max_delay = max_delay
        self.max_retries = max_retries
        self.timeout = timeout
        self.random = random.Random(seed)
        self.last_live_request = 0.0
        self.live_requests = 0
        self.cache_hits = 0

    def _cache_path(self, url: str) -> Path:
        digest = hashlib.sha256(url.encode("utf-8")).hexdigest()
        return self.cache_dir / f"{digest}.bin"

    def fetch(self, url: str, force: bool = False) -> bytes:
        cache_path = self._cache_path(url)
        if cache_path.exists() and not force:
            self.cache_hits += 1
            return cache_path.read_bytes()

        for attempt in range(self.max_retries + 1):
            elapsed = time.monotonic() - self.last_live_request
            required = self.random.uniform(self.min_delay, self.max_delay)
            if elapsed < required:
                time.sleep(required - elapsed)
            request = Request(
                url,
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/126.0 Safari/537.36"
                    ),
                    "Accept": "text/html,application/json;q=0.9,*/*;q=0.8",
                    "Accept-Language": "zh-CN,zh;q=0.9",
                },
            )
            try:
                self.last_live_request = time.monotonic()
                with urlopen(request, timeout=self.timeout) as response:
                    payload = response.read()
                self.live_requests += 1
                atomic_write_bytes(cache_path, payload)
                return payload
            except HTTPError as error:
                retryable = error.code in {403, 408, 429} or 500 <= error.code < 600
                if not retryable or attempt >= self.max_retries:
                    raise
                wait_seconds = min(240.0, 15.0 * (2**attempt)) + self.random.uniform(0, 5)
                print(f"HTTP {error.code}，{wait_seconds:.1f}秒后重试：{url}", flush=True)
                time.sleep(wait_seconds)
            except (URLError, TimeoutError, ConnectionError) as error:
                if attempt >= self.max_retries:
                    raise
                wait_seconds = min(240.0, 15.0 * (2**attempt)) + self.random.uniform(0, 5)
                print(f"网络错误 {error!s}，{wait_seconds:.1f}秒后重试：{url}", flush=True)
                time.sleep(wait_seconds)
        raise RuntimeError(f"无法获取：{url}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="采集并筛选作文网高中母语参照作文。")
    parser.add_argument("--learner-workbook", default=DEFAULT_LEARNER_WORKBOOK)
    parser.add_argument("--cache-dir", default=DEFAULT_CACHE_DIR)
    parser.add_argument("--ori-output-dir", default=DEFAULT_ORI_DIR)
    parser.add_argument("--clean-output-dir", default=DEFAULT_CLEAN_DIR)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--manual-review", default=DEFAULT_MANUAL_REVIEW)
    parser.add_argument("--target-per-group", type=int, default=DEFAULT_TARGET)
    parser.add_argument("--search-pages", type=int, default=4)
    parser.add_argument("--rare-search-pages", type=int, default=8)
    parser.add_argument("--middle-school-search-pages", type=int, default=8)
    parser.add_argument("--middle-school-category-tail-pages", type=int, default=30)
    parser.add_argument("--middle-school-category-page-step", type=int, default=5)
    parser.add_argument("--skip-middle-school-backup", action="store_true")
    parser.add_argument("--category-tail-pages", type=int, default=60)
    parser.add_argument("--category-page-step", type=int, default=3)
    parser.add_argument("--argument-category-tail-pages", type=int, default=1000)
    parser.add_argument("--argument-category-page-step", type=int, default=1)
    parser.add_argument("--skip-category-discovery", action="store_true")
    parser.add_argument("--skip-library-discovery", action="store_true")
    parser.add_argument(
        "--library-max-pages",
        type=int,
        default=0,
        help="每个高中作文文库筛选页最多抓取的页数；0表示全部。",
    )
    parser.add_argument("--pool-multiplier", type=float, default=2.0)
    parser.add_argument("--min-delay", type=float, default=3.0)
    parser.add_argument("--max-delay", type=float, default=6.0)
    parser.add_argument("--max-retries", type=int, default=5)
    parser.add_argument("--preferred-start-year", type=int, default=2005)
    parser.add_argument("--preferred-end-year", type=int, default=2012)
    parser.add_argument("--min-han", type=int, default=200)
    parser.add_argument("--max-han", type=int, default=750)
    parser.add_argument("--max-han-tolerance", type=float, default=0.05)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--force-fetch", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.target_per_group <= 0:
        raise ValueError("--target-per-group 必须大于0")
    if args.search_pages <= 0:
        raise ValueError("--search-pages 必须大于0")
    if args.rare_search_pages < args.search_pages:
        raise ValueError("--rare-search-pages 不得小于 --search-pages")
    if args.middle_school_search_pages <= 0:
        raise ValueError("--middle-school-search-pages 必须大于0")
    if args.middle_school_category_tail_pages < 0 or args.middle_school_category_page_step <= 0:
        raise ValueError("初中栏目分页参数无效")
    if args.category_tail_pages < 0 or args.category_page_step <= 0:
        raise ValueError("栏目分页参数无效")
    if args.argument_category_tail_pages < 0 or args.argument_category_page_step <= 0:
        raise ValueError("议论文栏目分页参数无效")
    if args.library_max_pages < 0:
        raise ValueError("--library-max-pages 不得小于0")
    if args.pool_multiplier < 1:
        raise ValueError("--pool-multiplier 不得小于1")
    if args.min_delay < 0 or args.max_delay < args.min_delay:
        raise ValueError("请求间隔参数无效")
    if args.min_han <= 0 or args.max_han < args.min_han:
        raise ValueError("正文汉字范围无效")
    if not 0 <= args.max_han_tolerance <= 0.10:
        raise ValueError("--max-han-tolerance 必须位于0至0.10之间")


def atomic_write_bytes(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", delete=False) as handle:
        handle.write(content)
        temp_path = Path(handle.name)
    temp_path.replace(path)


def atomic_write_text(path: Path, content: str) -> None:
    atomic_write_bytes(path, content.encode("utf-8"))


def decode_page(payload: bytes) -> tuple[str, str]:
    head = payload[:4096].lower()
    declared = re.search(br"charset\s*=\s*['\"]?([a-z0-9_-]+)", head)
    candidates: list[str] = []
    if declared:
        encoding = declared.group(1).decode("ascii", errors="ignore")
        if encoding in {"gb2312", "gbk"}:
            encoding = "gb18030"
        candidates.append(encoding)
    candidates.extend(["utf-8", "gb18030"])
    for encoding in dict.fromkeys(candidates):
        try:
            return payload.decode(encoding), encoding
        except (UnicodeDecodeError, LookupError):
            continue
    return payload.decode("utf-8", errors="replace"), "utf-8-replace"


def collapse_space(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(value)).strip()


def clean_source_label(value: str) -> str:
    label = collapse_space(value)
    label = re.sub(r"^来源[:：]\s*", "", label)
    for marker in (
        "E度作文网专稿",
        "E度网专稿",
        "作文网专稿",
        "未经允许",
        "本文系本站用户原创文章",
        "中学生写作指导",
        "中小学写作指导",
    ):
        if marker in label:
            label = label.split(marker, 1)[0].strip()
    return re.sub(r"[|｜/\s]+$", "", label).strip()


def normalize_title(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value)
    value = html.unescape(value)
    value = re.sub(r"^(?:高一|高二|高三)(?:年级)?(?:写人|叙事|议论)?作文[:：]\s*", "", value)
    value = re.sub(r"_\d{2,4}字\s*$", "", value)
    return collapse_space(value)


def base_title(value: str) -> str:
    result = normalize_title(value)
    result = SERIES_SUFFIX_RE.sub("", result)
    result = re.sub(r"[\s\W_]+", "", result)
    return result


def nominal_length(value: str) -> int | None:
    match = re.search(r"_(\d{2,4})字", value)
    return int(match.group(1)) if match else None


def normalize_url(value: str) -> str:
    value = value.strip()
    if value.startswith("/"):
        return f"https://www.zuowen.com{value}"
    parsed = urlparse(value)
    if parsed.hostname in {"www.zuowen.com", "zuowen.com", "m.zuowen.com"} and parsed.path:
        return f"https://www.zuowen.com{parsed.path}"
    return value.replace("http://", "https://", 1)


def is_zuowen_article_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.hostname in {"www.zuowen.com", "zuowen.com"} and bool(ARTICLE_ID_RE.search(parsed.path))


def nonoverlapping_term_occurrences(text: str, terms: Sequence[str]) -> int:
    occupied: list[tuple[int, int]] = []
    for term in sorted(set(terms), key=lambda value: (-len(value), value)):
        for match in re.finditer(re.escape(term), text):
            span = match.span()
            if any(span[0] < end and start < span[1] for start, end in occupied):
                continue
            occupied.append(span)
    return len(occupied)


def classify_topic(text: str, rule: TopicRule) -> tuple[str, float]:
    if "\n" in text:
        title, body = text.split("\n", 1)
        focus = f"{title} {body[:260]}"
    else:
        title = text
        focus = text
    title_compact = re.sub(r"\s+", "", title)
    compact = re.sub(r"\s+", "", focus)
    full_compact = re.sub(r"\s+", "", text)
    exact_hits = sum(term in compact for term in rule.exact_terms)
    near_hits = sum(term in compact for term in rule.near_terms)
    extended_hits = sum(term in compact for term in rule.extended_terms)
    broad_hits = sum(term in compact for term in rule.broad_terms)
    context_hits = sum(term in compact for term in rule.context_terms)
    all_topic_terms = (*rule.exact_terms, *rule.near_terms, *rule.extended_terms)
    full_occurrences = nonoverlapping_term_occurrences(full_compact, all_topic_terms)
    full_context_occurrences = nonoverlapping_term_occurrences(full_compact, rule.context_terms)
    broad_occurrences = nonoverlapping_term_occurrences(full_compact, rule.broad_terms)
    title_exact = any(term in title_compact for term in rule.exact_terms)
    title_near = any(term in title_compact for term in rule.near_terms)
    title_extended = any(term in title_compact for term in rule.extended_terms)
    title_broad = any(term in title_compact for term in rule.broad_terms)
    if exact_hits and (title_exact or full_occurrences >= 2) and (
        context_hits or rule.native_code in {"NJ2", "NY1", "NY2"}
    ):
        return "精确", 100 + exact_hits * 8 + context_hits + min(full_occurrences, 8)
    if near_hits and (title_near or full_occurrences >= 3) and (context_hits or extended_hits):
        return "近似", 70 + near_hits * 6 + context_hits + min(extended_hits, 4) + min(full_occurrences, 8)
    if extended_hits and (title_extended or full_occurrences >= 3) and (
        context_hits >= 2 or title_extended and full_context_occurrences >= 2
    ):
        return (
            "扩展",
            40
            + min(extended_hits, 5) * 3
            + context_hits
            + min(full_occurrences, 8)
            + min(full_context_occurrences, 5),
        )
    if broad_hits and (title_broad or broad_occurrences >= 3) and (
        context_hits >= 2 or title_broad and full_context_occurrences >= 2
    ):
        return (
            "宽泛",
            20
            + min(broad_hits, 5) * 2
            + context_hits
            + min(broad_occurrences, 8)
            + min(full_context_occurrences, 5),
        )
    return "不匹配", exact_hits * 8 + near_hits * 5 + extended_hits + broad_hits + context_hits


def parse_search_payload(payload: bytes, native_code: str, query: str) -> list[SearchCandidate]:
    text, _ = decode_page(payload)
    document = json.loads(text)
    if document.get("code") != 200:
        raise ValueError(f"搜索接口失败：{document}")
    rule = TOPIC_RULES[native_code]
    candidates: list[SearchCandidate] = []
    for item in document.get("data", {}).get("items", []):
        if item.get("data_type") != "article":
            continue
        url = normalize_url(str(item.get("target_url", "")))
        if not is_zuowen_article_url(url):
            continue
        raw_title = str(item.get("title", ""))
        title = normalize_title(raw_title)
        combined = f"{title} {collapse_space(str(item.get('content', '')))}"
        tier, score = classify_topic(combined, rule)
        candidates.append(
            SearchCandidate(
                native_code=native_code,
                url=url,
                title=title,
                search_date=str(item.get("ctime", "")),
                query=query,
                source_label=clean_source_label(str(item.get("copy_from", ""))),
                nominal_length=nominal_length(raw_title),
                pre_tier=tier,
                pre_score=score,
                verified_high_school_hint=bool(
                    re.search(r"高[一二三](?:年级)?(?:写人|叙事|议论|作文)", re.sub(r"<[^>]+>", "", raw_title))
                ),
            )
        )
    return candidates


def discover_candidates(
    client: SlowHttpClient,
    search_pages: int,
    force_fetch: bool,
    rare_search_pages: int | None = None,
) -> dict[str, list[SearchCandidate]]:
    result: dict[str, dict[str, SearchCandidate]] = {code: {} for code in TOPIC_RULES}
    for native_code, rule in TOPIC_RULES.items():
        print(f"发现 {native_code} 候选……", flush=True)
        page_count = (
            rare_search_pages
            if rare_search_pages is not None and native_code in {"NY1", "NY2"}
            else search_pages
        )
        for query in rule.queries:
            for page in range(1, page_count + 1):
                url = f"{SEARCH_ENDPOINT}?{urlencode({'site': 'zuowen', 'keyword': query, 'page': page})}"
                try:
                    payload = client.fetch(url, force=force_fetch)
                    page_candidates = parse_search_payload(payload, native_code, query)
                except Exception as error:
                    print(f"  搜索失败 {query} p{page}: {error}", file=sys.stderr, flush=True)
                    break
                if not page_candidates:
                    break
                for candidate in page_candidates:
                    existing = result[native_code].get(candidate.url)
                    if existing is None or candidate.pre_score > existing.pre_score:
                        result[native_code][candidate.url] = candidate
        print(f"  搜索去重后 {len(result[native_code])} 个URL", flush=True)
    return {code: list(values.values()) for code, values in result.items()}


def parse_listing_page(payload: bytes, native_codes: Sequence[str], source_url: str) -> list[SearchCandidate]:
    text, _ = decode_page(payload)
    block_pattern = re.compile(
        r'<div\s+class="artbox_l"[^>]*>.*?'
        r'<div\s+class="artbox_l_t"[^>]*>.*?'
        r'<a\s+[^>]*href="([^"]+)"[^>]*title="([^"]+)"[^>]*>.*?</a>\s*'
        r'<span>([^<]+)</span>.*?'
        r'<div\s+class="artbox_l_c"[^>]*>.*?<a(?:\s+[^>]*)?>(.*?)</a>',
        flags=re.I | re.S,
    )
    candidates: list[SearchCandidate] = []
    for url, raw_title, published, raw_excerpt in block_pattern.findall(text):
        normalized_url = normalize_url(url)
        if not is_zuowen_article_url(normalized_url):
            continue
        title = normalize_title(raw_title)
        excerpt = collapse_space(re.sub(r"<[^>]+>", "", raw_excerpt))
        for native_code in native_codes:
            rule = TOPIC_RULES[native_code]
            tier, score = classify_topic(f"{title} {excerpt}", rule)
            if tier == "不匹配":
                continue
            candidates.append(
                SearchCandidate(
                    native_code=native_code,
                    url=normalized_url,
                    title=title,
                    search_date=collapse_space(published),
                    query=f"高中栏目:{source_url}",
                    source_label="作文网高中年级栏目",
                    nominal_length=nominal_length(raw_title),
                    pre_tier=tier,
                    pre_score=score,
                    verified_high_school_hint=True,
                )
            )
    return candidates


def listing_max_page(payload: bytes) -> int:
    text, _ = decode_page(payload)
    values = [int(value) for value in re.findall(r"index_(\d+)\.shtml", text)]
    return max(values, default=1)


def category_page_url(base_url: str, page: int) -> str:
    if page <= 1:
        return base_url
    return f"{base_url.rstrip('/')}/index_{page}.shtml"


def discover_category_candidates(
    client: SlowHttpClient,
    tail_pages: int,
    page_step: int,
    argument_tail_pages: int,
    argument_page_step: int,
    force_fetch: bool,
) -> dict[str, list[SearchCandidate]]:
    result: dict[str, dict[str, SearchCandidate]] = {code: {} for code in TOPIC_RULES}
    for native_codes, base_url in CATEGORY_GROUP_URLS:
        is_argument_source = any(code in {"NY1", "NY2"} for code in native_codes)
        source_tail_pages = argument_tail_pages if is_argument_source else tail_pages
        source_page_step = argument_page_step if is_argument_source else page_step
        try:
            first_payload = client.fetch(base_url, force=force_fetch)
        except Exception as error:
            print(f"栏目首页失败 {base_url}: {error}", file=sys.stderr, flush=True)
            continue
        max_page = listing_max_page(first_payload)
        start_page = max(1, max_page - source_tail_pages)
        pages = list(range(start_page, max_page + 1, source_page_step))
        if max_page not in pages:
            pages.append(max_page)
        if 1 not in pages:
            pages.insert(0, 1)
        print(
            f"栏目发现 {','.join(native_codes)}：{base_url}，"
            f"页码1及{start_page}-{max_page}/{source_page_step}",
            flush=True,
        )
        for page in pages:
            try:
                payload = first_payload if page == 1 else client.fetch(
                    category_page_url(base_url, page), force=force_fetch
                )
                rows = parse_listing_page(payload, native_codes, base_url)
            except Exception as error:
                print(f"  栏目页{page}失败：{error}", file=sys.stderr, flush=True)
                continue
            for candidate in rows:
                existing = result[candidate.native_code].get(candidate.url)
                if existing is None or candidate.pre_score > existing.pre_score:
                    result[candidate.native_code][candidate.url] = candidate
        print(
            "  当前相关URL "
            + ", ".join(f"{code}={len(result[code])}" for code in native_codes),
            flush=True,
        )
    return {code: list(values.values()) for code, values in result.items()}


def discover_middle_school_ny2_candidates(
    client: SlowHttpClient,
    search_pages: int,
    category_tail_pages: int,
    category_page_step: int,
    force_fetch: bool,
) -> list[SearchCandidate]:
    result: dict[str, SearchCandidate] = {}
    for query in MIDDLE_SCHOOL_NY2_QUERIES:
        for page in range(1, search_pages + 1):
            url = f"{SEARCH_ENDPOINT}?{urlencode({'site': 'zuowen', 'keyword': query, 'page': page})}"
            try:
                rows = parse_search_payload(client.fetch(url, force=force_fetch), "NY2", query)
            except Exception as error:
                print(f"  初中备选搜索失败 {query} p{page}: {error}", file=sys.stderr, flush=True)
                break
            if not rows:
                break
            for candidate in rows:
                candidate.middle_school_backup = True
                candidate.query = f"初中备选:{candidate.query}"
                existing = result.get(candidate.url)
                if existing is None or candidate.pre_score > existing.pre_score:
                    result[candidate.url] = candidate

    for base_url in MIDDLE_SCHOOL_NY2_CATEGORY_URLS:
        try:
            first_payload = client.fetch(base_url, force=force_fetch)
        except Exception as error:
            print(f"初中栏目首页失败 {base_url}: {error}", file=sys.stderr, flush=True)
            continue
        max_page = listing_max_page(first_payload)
        start_page = max(1, max_page - category_tail_pages)
        pages = list(range(start_page, max_page + 1, category_page_step))
        if max_page not in pages:
            pages.append(max_page)
        if 1 not in pages:
            pages.insert(0, 1)
        for page in pages:
            try:
                payload = first_payload if page == 1 else client.fetch(
                    category_page_url(base_url, page), force=force_fetch
                )
                rows = parse_listing_page(payload, ("NY2",), base_url)
            except Exception as error:
                print(f"  初中栏目页{page}失败：{error}", file=sys.stderr, flush=True)
                continue
            for candidate in rows:
                candidate.middle_school_backup = True
                candidate.verified_high_school_hint = False
                candidate.query = f"初中备选栏目:{base_url}"
                candidate.source_label = "作文网初中年级栏目"
                existing = result.get(candidate.url)
                if existing is None or candidate.pre_score > existing.pre_score:
                    result[candidate.url] = candidate
    print(f"初中NY2备选去重URL：{len(result)}", flush=True)
    return list(result.values())


def parse_library_listing_page(
    payload: bytes,
    native_codes: Sequence[str],
    source_url: str,
) -> list[SearchCandidate]:
    text, _ = decode_page(payload)
    blocks = re.split(r'<div\s+class="list">', text, flags=re.I)[1:]
    blocks = [block.split('<div class="clear"></div>', 1)[0] for block in blocks]
    grade_map = {"10": "高一", "11": "高二", "12": "高三"}
    genre_map = {"1": "写人作文", "2": "叙事作文", "6": "话题作文", "9": "议论文"}
    candidates: list[SearchCandidate] = []
    for block in blocks:
        link_match = re.search(
            r'<b>\s*<a\s+[^>]*href="([^"]+)"[^>]*>(.*?)</a>\s*</b>',
            block,
            flags=re.I | re.S,
        )
        excerpt_match = re.search(
            r'<p\s+class="article"[^>]*>(.*?)</p>',
            block,
            flags=re.I | re.S,
        )
        tags_match = re.search(
            r'<div\s+class="tags"[^>]*>(.*?)</div>',
            block,
            flags=re.I | re.S,
        )
        if not link_match or not excerpt_match or not tags_match:
            continue
        normalized_url = normalize_url(link_match.group(1))
        if not is_zuowen_article_url(normalized_url):
            continue
        raw_title = collapse_space(re.sub(r"<[^>]+>", "", link_match.group(2)))
        title = normalize_title(raw_title)
        excerpt = collapse_space(re.sub(r"<[^>]+>", "", excerpt_match.group(1)))
        tags = tags_match.group(1)
        grade_match = re.search(r"/wk/gz-(10|11|12)-", tags)
        genre_codes = re.findall(r"/wk/gz-\d+-(\d+)-", tags)
        genre_code = next((value for value in genre_codes if value in genre_map), "")
        if not grade_match or not genre_code:
            continue
        published_match = re.search(
            r'class="date"[^>]*>\s*(?:投稿时间[:：]\s*)?([^<]+)</span>',
            block,
            flags=re.I,
        )
        grade = grade_map[grade_match.group(1)]
        genre = genre_map[genre_code]
        for native_code in native_codes:
            rule = TOPIC_RULES[native_code]
            tier, score = classify_topic(f"{title}\n{excerpt}", rule)
            if tier == "不匹配":
                continue
            candidates.append(
                SearchCandidate(
                    native_code=native_code,
                    url=normalized_url,
                    title=title,
                    search_date=collapse_space(published_match.group(1)) if published_match else "",
                    query=f"高中作文文库:{source_url}",
                    source_label="作文网高中作文文库",
                    nominal_length=nominal_length(raw_title),
                    pre_tier=tier,
                    pre_score=score,
                    verified_high_school_hint=True,
                    library_verified=True,
                    verified_grade=grade,
                    verified_genre=genre,
                )
            )
    return candidates


def library_page_url(genre_code: int, length_code: int, page: int, sort_code: int = 2) -> str:
    page_code = 0 if page <= 1 else page
    return f"https://www.zuowen.com/wk/gz-0-{genre_code}-{length_code}-{page_code}-{sort_code}.html"


def library_max_page(payload: bytes) -> int:
    text, _ = decode_page(payload)
    values = [
        int(value)
        for value in re.findall(r"/wk/gz-\d+-\d+-\d+-(\d+)-[12]\.html", text)
        if int(value) > 0
    ]
    return max(values, default=1)


def discover_library_candidates(
    client: SlowHttpClient,
    max_pages: int,
    force_fetch: bool,
) -> dict[str, list[SearchCandidate]]:
    result: dict[str, dict[str, SearchCandidate]] = {code: {} for code in TOPIC_RULES}
    for native_codes, genre_code, genre_label in LIBRARY_GROUP_GENRES:
        print(f"文库发现 {','.join(native_codes)}：{genre_label}", flush=True)
        for length_code in LIBRARY_LENGTH_CODES:
            first_url = library_page_url(genre_code, length_code, 1)
            try:
                first_payload = client.fetch(first_url, force=force_fetch)
            except Exception as error:
                print(f"  文库首页失败 {first_url}: {error}", file=sys.stderr, flush=True)
                continue
            page_count = library_max_page(first_payload)
            if max_pages:
                page_count = min(page_count, max_pages)
            for page in range(1, page_count + 1):
                try:
                    payload = first_payload if page == 1 else client.fetch(
                        library_page_url(genre_code, length_code, page),
                        force=force_fetch,
                    )
                    rows = parse_library_listing_page(payload, native_codes, first_url)
                except Exception as error:
                    print(f"  文库页{page}失败：{error}", file=sys.stderr, flush=True)
                    continue
                for candidate in rows:
                    existing = result[candidate.native_code].get(candidate.url)
                    if existing is None or candidate.pre_score > existing.pre_score:
                        result[candidate.native_code][candidate.url] = candidate
        print(
            "  当前文库相关URL "
            + ", ".join(f"{code}={len(result[code])}" for code in native_codes),
            flush=True,
        )
    return {code: list(values.values()) for code, values in result.items()}


def merge_discovered(
    first: dict[str, list[SearchCandidate]],
    second: dict[str, list[SearchCandidate]],
) -> dict[str, list[SearchCandidate]]:
    merged: dict[str, dict[str, SearchCandidate]] = {code: {} for code in TOPIC_RULES}
    for source in (first, second):
        for code, rows in source.items():
            for candidate in rows:
                existing = merged[code].get(candidate.url)
                if existing is None:
                    merged[code][candidate.url] = candidate
                elif existing.middle_school_backup and not candidate.middle_school_backup:
                    merged[code][candidate.url] = candidate
                elif candidate.middle_school_backup and not existing.middle_school_backup:
                    continue
                elif candidate.library_verified and not existing.library_verified:
                    merged[code][candidate.url] = candidate
                elif existing.library_verified and not candidate.library_verified:
                    continue
                elif candidate.verified_high_school_hint and not existing.verified_high_school_hint:
                    merged[code][candidate.url] = candidate
                elif candidate.pre_score > existing.pre_score:
                    merged[code][candidate.url] = candidate
    return {code: list(values.values()) for code, values in merged.items()}


def clean_body(raw_text: str, title: str) -> str:
    text = raw_text.replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    text = html.unescape(text)
    lines: list[str] = []
    for line in text.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        if not line:
            if lines and lines[-1] != "":
                lines.append("")
            continue
        if normalize_title(line) == normalize_title(title):
            continue
        for pattern in BOILERPLATE_PATTERNS:
            line = pattern.sub("", line).strip()
        if not line:
            continue
        if re.match(r"^(?:高|初)[一二三]\s*[:：].{0,80}$", line):
            continue
        if re.match(r"^(?:高一|高二|高三|高中|初一|初二|初三|初中).{0,40}(?:作者|学生)[:：]", line):
            continue
        lines.append(line)
    while lines and not lines[-1]:
        lines.pop()
    tail_author = re.compile(
        r"^.{0,35}(?:中学|学校|一中|二中|三中|四中|五中|六中|七中|八中|九中|十中)"
        r".{0,25}(?:高[一二三\d]|中专|年级|班).{0,25}$"
    )
    while lines and tail_author.match(lines[-1]):
        lines.pop()
        while lines and not lines[-1]:
            lines.pop()
    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def load_manual_reviews(path: Path) -> dict[tuple[str, str], tuple[str, str]]:
    if not path.exists():
        return {}
    result: dict[tuple[str, str], tuple[str, str]] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            native_code = collapse_space(row.get("母语代码", ""))
            url = normalize_url(row.get("URL", ""))
            decision = collapse_space(row.get("决定", ""))
            reason = collapse_space(row.get("原因", ""))
            if native_code not in TOPIC_RULES or not is_zuowen_article_url(url):
                raise ValueError(f"人工审核表存在无效代码或URL：{row}")
            if decision not in {"排除", "保留"}:
                raise ValueError(f"人工审核决定必须是“排除”或“保留”：{row}")
            result[(native_code, url)] = (decision, reason)
    return result


def parse_article(payload: bytes, url: str, fallback_source: str = "") -> ParsedArticle:
    text, _ = decode_page(payload)
    parser = ZuowenHTMLParser()
    parser.feed(text)
    title = normalize_title("".join(parser.title_parts))
    breadcrumb = collapse_space("".join(parser.path_parts))
    raw_text = "".join(parser.body_parts).strip()
    if not title:
        title_match = re.search(r"<title>\s*(.*?)\s*</title>", text, flags=re.I | re.S)
        title = normalize_title(title_match.group(1) if title_match else "")
    if not raw_text:
        raise ValueError("未找到 con_content 正文")
    meta = collapse_space("".join(parser.meta_parts))
    date_match = DATE_RE.search(meta) or DATE_RE.search(url)
    published_date = ""
    if date_match:
        published_date = f"{int(date_match.group(1)):04d}-{int(date_match.group(2)):02d}-{int(date_match.group(3)):02d}"
    source_match = re.search(r"来源[:：]\s*([^\s]+(?:\s*[^\s]+)?)", meta)
    source_label = clean_source_label(source_match.group(1) if source_match else fallback_source)
    grade_match = re.search(r"(?:高|初)[一二三]", breadcrumb + " " + title)
    tail_grade_match = re.search(r"(?:^|\n)\s*((?:高|初)[一二三])\s*[:：]", raw_text)
    grade = (
        grade_match.group(0)
        if grade_match
        else tail_grade_match.group(1)
        if tail_grade_match
        else "未标注"
    )
    genre_label = ""
    for label in (
        "写人作文",
        "叙事作文",
        "议论文",
        "议论作文",
        "话题作文",
        "叙事散文",
        "议论散文",
    ):
        if label in breadcrumb or label in title:
            genre_label = label
            break
    clean_text = clean_body(raw_text, title)
    han_count = len(HAN_RE.findall(clean_text))
    article_id_match = ART_ID_RE.search(text) or ARTICLE_ID_RE.search(url)
    article_id = article_id_match.group(1) if article_id_match else hashlib.sha1(url.encode()).hexdigest()[:16]
    normalized_for_hash = re.sub(r"\s+", "", clean_text)
    return ParsedArticle(
        url=url,
        article_id=article_id,
        title=title,
        breadcrumb=breadcrumb,
        grade=grade,
        genre_label=genre_label,
        published_date=published_date,
        source_label=source_label,
        raw_text=raw_text,
        clean_text=clean_text,
        han_count=han_count,
        text_hash=hashlib.sha256(normalized_for_hash.encode("utf-8")).hexdigest(),
    )


def argumentative_signal_count(text: str) -> int:
    markers = (
        "我认为",
        "在我看来",
        "首先",
        "其次",
        "因此",
        "所以",
        "然而",
        "但是",
        "可见",
        "总之",
        "由此",
        "不仅",
        "更应该",
        "不应该",
        "应该",
    )
    return sum(marker in text for marker in markers)


def is_high_school_source(breadcrumb: str) -> bool:
    return "高中作文" in breadcrumb or ("单元作文" in breadcrumb and "高中语文" in breadcrumb)


def is_middle_school_source(breadcrumb: str) -> bool:
    return "初中作文" in breadcrumb or ("单元作文" in breadcrumb and "初中语文" in breadcrumb)


def validate_article(
    article: ParsedArticle,
    rule: TopicRule,
    min_han: int,
    max_han: int,
    allow_middle_school: bool = False,
) -> tuple[bool, str]:
    combined_header = f"{article.breadcrumb} {article.title}"
    is_high_school = is_high_school_source(article.breadcrumb)
    is_middle_school = is_middle_school_source(article.breadcrumb)
    if not is_high_school and not (allow_middle_school and is_middle_school):
        return False, "面包屑不属于允许的高中或初中备选作文"
    excluded = next((term for term in EXCLUDED_TERMS if term in combined_header), None)
    if excluded:
        return False, f"排除来源类别：{excluded}"
    if rule.genre == "记叙文":
        explicit_narrative = any(
            term in combined_header for term in ("叙事作文", "写人作文", "叙事", "写人")
        )
        high_school_unit_narrative = (
            is_high_school
            and "单元作文" in article.breadcrumb
            and classify_topic(f"{article.title}\n{article.clean_text}", rule)[0] != "不匹配"
        )
        if not explicit_narrative and not high_school_unit_narrative:
            return False, "体裁不是叙事或写人"
    else:
        explicit_argument = any(term in combined_header for term in ("议论文", "议论作文", "议论"))
        inferred_argument = (
            (is_high_school or allow_middle_school and is_middle_school)
            and (article.genre_label == "话题作文" or "单元作文" in article.breadcrumb)
            and argumentative_signal_count(article.clean_text) >= 2
        )
        if not explicit_argument and not inferred_argument:
            return False, "体裁不是议论文"
    if article.han_count < min_han or article.han_count > max_han:
        return False, f"正文汉字数超出{min_han}-{max_han}"
    if len(article.clean_text.splitlines()) == 0:
        return False, "正文为空"
    return True, ""


def candidate_sort_key(candidate: SearchCandidate, preferred_start: int, preferred_end: int) -> tuple[Any, ...]:
    year = int(candidate.search_date[:4]) if candidate.search_date[:4].isdigit() else 9999
    period_penalty = 0 if preferred_start <= year <= preferred_end else abs(year - preferred_end) + 10
    tier_rank = {"精确": 0, "近似": 1, "扩展": 2, "宽泛": 3, "不匹配": 4}[candidate.pre_tier]
    return (
        1 if candidate.middle_school_backup else 0,
        tier_rank,
        0 if candidate.verified_high_school_hint else 1,
        period_penalty,
        -candidate.pre_score,
        candidate.url,
    )


def build_eligible_pools(
    client: SlowHttpClient,
    discovered: dict[str, list[SearchCandidate]],
    manual_reviews: dict[tuple[str, str], tuple[str, str]],
    target_per_group: int,
    pool_multiplier: float,
    preferred_start: int,
    preferred_end: int,
    min_han: int,
    max_han: int,
    force_fetch: bool,
) -> tuple[dict[str, list[CandidateAudit]], list[CandidateAudit]]:
    desired_pool = max(target_per_group, math.ceil(target_per_group * pool_multiplier))
    pools: dict[str, list[CandidateAudit]] = {code: [] for code in TOPIC_RULES}
    audit: list[CandidateAudit] = []
    parsed_by_url: dict[str, ParsedArticle] = {}
    shared_cap: int | None = None
    scan_order = ["NY2", "NY1", "NJ2", "NJ1"]
    print(f"候选核验顺序：{' -> '.join(scan_order)}", flush=True)
    for native_code in scan_order:
        rule = TOPIC_RULES[native_code]
        group_desired = desired_pool
        if shared_cap is not None and shared_cap < target_per_group:
            group_desired = min(
                desired_pool,
                max(6, shared_cap, math.ceil(shared_cap * pool_multiplier)),
            )
        ordered = sorted(
            discovered[native_code],
            key=lambda item: candidate_sort_key(item, preferred_start, preferred_end),
        )
        for index, search_candidate in enumerate(ordered, start=1):
            if len(pools[native_code]) >= group_desired:
                break
            record = CandidateAudit(
                native_code=native_code,
                learner_code=rule.learner_code,
                url=search_candidate.url,
                title=search_candidate.title,
                published_date=search_candidate.search_date,
                source_label=search_candidate.source_label,
                query=search_candidate.query,
                nominal_length=search_candidate.nominal_length,
            )
            if search_candidate.pre_tier == "不匹配":
                record.status = "拒绝"
                record.reason = "搜索标题与摘要主题不匹配，未请求正文"
                audit.append(record)
                if index % 10 == 0:
                    print(
                        f"  {native_code} {index}/{len(ordered)} "
                        f"候选={len(pools[native_code])}/{group_desired} "
                        f"{record.status} {record.title[:24]}",
                        flush=True,
                    )
                continue
            try:
                article = parsed_by_url.get(search_candidate.url)
                if article is None:
                    payload = client.fetch(search_candidate.url, force=force_fetch)
                    article = parse_article(payload, search_candidate.url, search_candidate.source_label)
                    parsed_by_url[search_candidate.url] = article
                effective_article = article
                if search_candidate.library_verified:
                    source_label = "作文网高中作文文库"
                    if article.source_label and article.source_label != source_label:
                        source_label += f"（{article.source_label}）"
                    effective_article = replace(
                        article,
                        breadcrumb=(
                            f"作文 > 高中作文 > 高中作文文库 > "
                            f"{search_candidate.verified_grade} > {search_candidate.verified_genre}"
                        ),
                        grade=search_candidate.verified_grade or article.grade,
                        genre_label=search_candidate.verified_genre or article.genre_label,
                        source_label=source_label,
                    )
                record.article = effective_article
                record.article_id = effective_article.article_id
                record.title = effective_article.title
                record.published_date = effective_article.published_date or search_candidate.search_date
                record.grade = effective_article.grade
                record.school_stage = (
                    "高中"
                    if is_high_school_source(effective_article.breadcrumb)
                    else "初中"
                    if is_middle_school_source(effective_article.breadcrumb)
                    else "未核验"
                )
                record.genre_label = effective_article.genre_label
                record.source_label = effective_article.source_label or search_candidate.source_label
                record.han_count = effective_article.han_count
                record.text_hash = effective_article.text_hash
                valid, reason = validate_article(
                    effective_article,
                    rule,
                    min_han,
                    max_han,
                    allow_middle_school=search_candidate.middle_school_backup,
                )
                tier, topic_score = classify_topic(
                    f"{effective_article.title}\n{effective_article.clean_text}", rule
                )
                record.topic_tier = tier
                record.topic_score = topic_score
                manual_decision, manual_reason = manual_reviews.get(
                    (native_code, search_candidate.url),
                    ("", ""),
                )
                if not valid:
                    record.status = "拒绝"
                    record.reason = reason
                elif manual_decision == "排除":
                    record.status = "拒绝"
                    record.reason = f"人工审核排除：{manual_reason or '未注明原因'}"
                elif tier == "不匹配":
                    record.status = "拒绝"
                    record.reason = "主题不匹配"
                else:
                    record.status = "候选"
                    pools[native_code].append(record)
            except Exception as error:
                record.status = "失败"
                record.reason = str(error)[:240]
            audit.append(record)
            if index % 10 == 0 or record.status == "候选":
                print(
                    f"  {native_code} {index}/{len(ordered)} 候选={len(pools[native_code])}/{group_desired} "
                    f"{record.status} {record.title[:24]}",
                    flush=True,
                )
        print(f"{native_code} 合格候选池：{len(pools[native_code])}", flush=True)
        group_cap = min(target_per_group, len(pools[native_code]))
        shared_cap = group_cap if shared_cap is None else min(shared_cap, group_cap)
    return pools, audit


def character_ngrams(text: str, n: int = 5) -> frozenset[str]:
    compact = re.sub(r"\s+", "", text)
    if len(compact) <= n:
        return frozenset({compact}) if compact else frozenset()
    return frozenset(compact[index : index + n] for index in range(len(compact) - n + 1))


def jaccard(first: frozenset[str], second: frozenset[str]) -> float:
    if not first and not second:
        return 1.0
    union = len(first | second)
    return len(first & second) / union if union else 0.0


def deduplicate_pool(pool: Sequence[CandidateAudit], threshold: float = 0.85) -> tuple[list[CandidateAudit], list[CandidateAudit]]:
    tier_rank = {"精确": 0, "近似": 1, "扩展": 2, "宽泛": 3, "": 4}
    ordered = sorted(
        pool,
        key=lambda item: (
            0 if item.school_stage == "高中" else 1,
            tier_rank.get(item.topic_tier, 3),
            0 if 2005 <= publication_year(item.published_date) <= 2012 else 1,
            -item.topic_score,
            item.han_count,
        ),
    )
    kept: list[CandidateAudit] = []
    rejected: list[CandidateAudit] = []
    hashes: set[str] = set()
    titles: set[str] = set()
    grams: list[frozenset[str]] = []
    for candidate in ordered:
        assert candidate.article is not None
        title_key = base_title(candidate.title)
        reason = ""
        if candidate.text_hash in hashes:
            reason = "正文哈希重复"
        elif title_key and title_key in titles:
            reason = "标准化题目重复"
        else:
            current = character_ngrams(candidate.article.clean_text)
            if any(jaccard(current, previous) >= threshold for previous in grams):
                reason = "正文近重复（五元组Jaccard>=0.85）"
        if reason:
            candidate.status = "拒绝"
            candidate.reason = reason
            rejected.append(candidate)
            continue
        kept.append(candidate)
        hashes.add(candidate.text_hash)
        if title_key:
            titles.add(title_key)
        grams.append(character_ngrams(candidate.article.clean_text))
    return kept, rejected


def deduplicate_across_groups(
    pools: dict[str, list[CandidateAudit]],
    threshold: float = 0.85,
) -> tuple[dict[str, list[CandidateAudit]], list[CandidateAudit]]:
    tier_rank = {"精确": 0, "近似": 1, "扩展": 2, "宽泛": 3, "": 4}
    ordered = sorted(
        (item for rows in pools.values() for item in rows),
        key=lambda item: (
            0 if item.school_stage == "高中" else 1,
            tier_rank.get(item.topic_tier, 4),
            0 if 2005 <= publication_year(item.published_date) <= 2012 else 1,
            -item.topic_score,
            item.native_code,
        ),
    )
    kept: dict[str, list[CandidateAudit]] = {code: [] for code in pools}
    accepted: list[tuple[CandidateAudit, frozenset[str]]] = []
    rejected: list[CandidateAudit] = []
    for candidate in ordered:
        assert candidate.article is not None
        current = character_ngrams(candidate.article.clean_text)
        duplicate_of: CandidateAudit | None = None
        duplicate_reason = ""
        for previous, previous_grams in accepted:
            if previous.native_code == candidate.native_code:
                continue
            if candidate.text_hash == previous.text_hash:
                duplicate_of = previous
                duplicate_reason = "正文哈希重复"
                break
            if jaccard(current, previous_grams) >= threshold:
                duplicate_of = previous
                duplicate_reason = "正文近重复（五元组Jaccard>=0.85）"
                break
        if duplicate_of is not None:
            candidate.status = "拒绝"
            candidate.reason = (
                f"跨组{duplicate_reason}，保留于{duplicate_of.native_code}：{duplicate_of.title}"
            )
            rejected.append(candidate)
            continue
        kept[candidate.native_code].append(candidate)
        accepted.append((candidate, current))
    return kept, rejected


def publication_year(value: str) -> int:
    return int(value[:4]) if len(value) >= 4 and value[:4].isdigit() else 9999


def read_learner_lengths(workbook_path: Path) -> dict[str, list[int]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("缺少 openpyxl，无法读取学习者宽表") from error
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["词性统计"] if "词性统计" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [str(cell.value or "") for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: headers.index(name) for name in ("篇名代码", "纯文本字数")}
    result: dict[str, list[int]] = {rule.learner_code: [] for rule in TOPIC_RULES.values()}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[index["篇名代码"]] or "")
        if code in result:
            result[code].append(int(row[index["纯文本字数"]]))
    workbook.close()
    for code, values in result.items():
        if len(values) != 155:
            raise ValueError(f"学习者组{code}应有155篇，实际{len(values)}")
    return result


def quantile_targets(values: Sequence[int], count: int) -> list[int]:
    ordered = sorted(values)
    if count == 1:
        return [round(statistics.median(ordered))]
    targets: list[int] = []
    for index in range(count):
        position = (index + 0.5) / count * (len(ordered) - 1)
        lower = math.floor(position)
        upper = math.ceil(position)
        fraction = position - lower
        targets.append(round(ordered[lower] * (1 - fraction) + ordered[upper] * fraction))
    return targets


def selection_cost(
    candidate: CandidateAudit,
    target: int,
    grade_counts: Counter[str],
    grade_quotas: dict[str, int],
    preferred_start: int,
    preferred_end: int,
) -> float:
    tier_penalty = {"精确": 0.0, "近似": 0.16, "扩展": 0.38, "宽泛": 0.68}[
        candidate.topic_tier
    ]
    year = publication_year(candidate.published_date)
    if preferred_start <= year <= preferred_end:
        year_penalty = 0.0
    elif year == 9999:
        year_penalty = 0.20
    else:
        year_penalty = min(0.35, 0.02 * min(abs(year - preferred_end), 15))
    grade = candidate.grade
    if grade == "未标注":
        grade_penalty = 0.16 + max(0, grade_counts[grade] - 9) * 2
    elif grade not in grade_quotas:
        grade_penalty = 0.12
    elif grade_counts[grade] >= grade_quotas.get(grade, 0):
        grade_penalty = 0.28 + (grade_counts[grade] - grade_quotas[grade]) * 0.08
    else:
        grade_penalty = 0.0
    stage_penalty = 0.0 if candidate.school_stage == "高中" else 0.55
    length_penalty = abs(candidate.han_count - target) / max(target, 1)
    return (
        length_penalty
        + tier_penalty
        + year_penalty
        + grade_penalty
        + stage_penalty
        - min(candidate.topic_score, 140) / 3000
    )


def select_group(
    pool: Sequence[CandidateAudit],
    learner_lengths: Sequence[int],
    count: int,
    preferred_start: int,
    preferred_end: int,
) -> list[CandidateAudit]:
    targets = quantile_targets(learner_lengths, count)
    remaining = list(pool)
    selected: list[CandidateAudit] = []
    grade_counts: Counter[str] = Counter()
    base = count // 3
    extras = count - base * 3
    grade_quotas = {
        "高一": base,
        "高二": base + (1 if extras >= 2 else 0),
        "高三": base + (1 if extras >= 1 else 0),
    }
    for target in targets:
        if not remaining:
            break
        slots_left = count - len(selected)
        high_school_remaining = [item for item in remaining if item.school_stage == "高中"]
        eligible = high_school_remaining if len(high_school_remaining) >= slots_left else remaining
        best = min(
            eligible,
            key=lambda item: selection_cost(
                item,
                target,
                grade_counts,
                grade_quotas,
                preferred_start,
                preferred_end,
            ),
        )
        setattr(best, "_target_han", target)
        selected.append(best)
        grade_counts[best.grade] += 1
        remaining.remove(best)
    return selected


def sample_standard_deviation(values: Sequence[int]) -> float:
    return statistics.stdev(values) if len(values) > 1 else 1.0


def standardized_mean_difference(first: Sequence[int], second: Sequence[int]) -> float:
    first_variance = statistics.variance(first) if len(first) > 1 else 0.0
    second_variance = statistics.variance(second) if len(second) > 1 else 0.0
    pooled = math.sqrt((first_variance + second_variance) / 2)
    return (statistics.mean(first) - statistics.mean(second)) / pooled if pooled > 0 else 0.0


def prune_generated_texts(base_dir: Path, expected_stems: dict[str, set[str]]) -> list[Path]:
    """Remove stale generated text files while leaving all other files untouched."""
    removed: list[Path] = []
    for native_code, allowed in expected_stems.items():
        code_dir = base_dir / native_code
        if not code_dir.is_dir():
            continue
        for path in code_dir.glob("*.txt"):
            if path.stem not in allowed:
                path.unlink()
                removed.append(path)
    return removed


def write_outputs(
    selected_by_code: dict[str, list[CandidateAudit]],
    audit: list[CandidateAudit],
    learner_lengths: dict[str, list[int]],
    ori_dir: Path,
    clean_dir: Path,
    output_dir: Path,
    preferred_start: int,
    preferred_end: int,
    nominal_min_han: int,
    nominal_max_han: int,
    dry_run: bool,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    selected_rows: list[dict[str, Any]] = []
    summary_groups: dict[str, Any] = {}
    for native_code, selected in selected_by_code.items():
        rule = TOPIC_RULES[native_code]
        native_lengths = [item.han_count for item in selected]
        learner_values = learner_lengths[rule.learner_code]
        learner_sd = sample_standard_deviation(learner_values)
        for sequence, item in enumerate(sorted(selected, key=lambda row: (row.han_count, row.url)), start=1):
            item.status = "入选"
            filename = f"{native_code}_{sequence:03d}"
            target_han = int(getattr(item, "_target_han", round(statistics.median(learner_values))))
            assert item.article is not None
            if not dry_run:
                atomic_write_text(ori_dir / native_code / f"{filename}.txt", item.article.raw_text.strip() + "\n")
                atomic_write_text(clean_dir / native_code / f"{filename}.txt", item.article.clean_text.strip() + "\n")
            selected_rows.append(
                {
                    "母语代码": native_code,
                    "对应学习者代码": rule.learner_code,
                    "对应学习者篇名": rule.learner_title,
                    "作文文件名": filename,
                    "网页文章ID": item.article_id,
                    "作文题目": item.title,
                    "体裁": rule.genre,
                    "网站体裁": item.genre_label,
                    "学段": item.school_stage,
                    "年级": item.grade,
                    "发布日期": item.published_date,
                    "来源类型": item.source_label,
                    "来源URL": item.url,
                    "主题匹配层级": item.topic_tier,
                    "主题匹配得分": item.topic_score,
                    "正文汉字数": item.han_count,
                    "目标汉字数": target_han,
                    "目标篇幅相对偏差": abs(item.han_count - target_han) / max(target_han, 1),
                    "目标篇幅标准差偏差": abs(item.han_count - target_han) / max(learner_sd, 1),
                    "篇幅范围扩展样本": not (
                        nominal_min_han <= item.han_count <= nominal_max_han
                    ),
                    "学段扩展样本": item.school_stage != "高中",
                    "年代扩展样本": not (
                        preferred_start <= publication_year(item.published_date) <= preferred_end
                    ),
                    "正文SHA256": item.text_hash,
                    "审核状态": "自动筛选通过，待全文人工复核",
                }
            )
        summary_groups[native_code] = {
            "selected": len(selected),
            "learner_code": rule.learner_code,
            "grade_counts": dict(Counter(item.grade for item in selected)),
            "stage_counts": dict(Counter(item.school_stage for item in selected)),
            "tier_counts": dict(Counter(item.topic_tier for item in selected)),
            "year_range": [min(map(publication_year, (item.published_date for item in selected))), max(map(publication_year, (item.published_date for item in selected)))],
            "native_han_mean": statistics.mean(native_lengths) if native_lengths else None,
            "native_han_median": statistics.median(native_lengths) if native_lengths else None,
            "learner_han_mean": statistics.mean(learner_values),
            "learner_han_median": statistics.median(learner_values),
            "length_smd": standardized_mean_difference(native_lengths, learner_values) if native_lengths else None,
        }

    selected_rows.sort(key=lambda row: (row["母语代码"], row["作文文件名"]))
    if not dry_run:
        expected_stems = {
            native_code: {
                row["作文文件名"]
                for row in selected_rows
                if row["母语代码"] == native_code
            }
            for native_code in TOPIC_RULES
        }
        removed = prune_generated_texts(ori_dir, expected_stems)
        removed.extend(prune_generated_texts(clean_dir, expected_stems))
        if removed:
            print(f"已清理旧版入选文本：{len(removed)} 个", flush=True)
    serializable_audit = []
    selected_urls = {row["来源URL"] for row in selected_rows}
    for item in audit:
        if item.url in selected_urls:
            item.status = "入选"
        serializable_audit.append(
            {
                key: value
                for key, value in asdict(item).items()
                if key != "article" and key not in {"raw_text", "clean_text"}
            }
        )

    payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_site": "https://www.zuowen.com/gaozhong/",
        "source_agreement": "http://www.zuowen.com/help/agreement/",
        "corpus_label": "公开网络母语参照语料（作者身份与编辑情况未独立核验）",
        "selected": selected_rows,
        "summary": summary_groups,
    }
    if not dry_run:
        atomic_write_text(output_dir / "selected_samples.json", json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
        atomic_write_text(output_dir / "selection_summary.json", json.dumps(payload["summary"], ensure_ascii=False, indent=2) + "\n")
        audit_path = output_dir / "candidate_audit.csv"
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = list(serializable_audit[0]) if serializable_audit else ["native_code", "url", "status", "reason"]
        with NamedTemporaryFile(
            mode="w",
            encoding="utf-8-sig",
            newline="",
            dir=audit_path.parent,
            prefix=f".{audit_path.name}.",
            delete=False,
        ) as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(serializable_audit)
            temp_path = Path(handle.name)
        temp_path.replace(audit_path)
    return payload


def main() -> int:
    args = parse_args()
    validate_args(args)
    learner_workbook = Path(args.learner_workbook).resolve()
    cache_dir = Path(args.cache_dir).resolve()
    ori_dir = Path(args.ori_output_dir).resolve()
    clean_dir = Path(args.clean_output_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    manual_reviews = load_manual_reviews(Path(args.manual_review).resolve())
    learner_lengths = read_learner_lengths(learner_workbook)
    client = SlowHttpClient(
        cache_dir=cache_dir,
        min_delay=args.min_delay,
        max_delay=args.max_delay,
        max_retries=args.max_retries,
        seed=args.seed,
    )
    discovered = discover_candidates(
        client,
        args.search_pages,
        args.force_fetch,
        rare_search_pages=args.rare_search_pages,
    )
    if not args.skip_category_discovery:
        category_discovered = discover_category_candidates(
            client,
            tail_pages=args.category_tail_pages,
            page_step=args.category_page_step,
            argument_tail_pages=args.argument_category_tail_pages,
            argument_page_step=args.argument_category_page_step,
            force_fetch=args.force_fetch,
        )
        discovered = merge_discovered(category_discovered, discovered)
        print(
            "合并候选：" + ", ".join(f"{code}={len(rows)}" for code, rows in discovered.items()),
            flush=True,
        )
    if not args.skip_library_discovery:
        library_discovered = discover_library_candidates(
            client,
            max_pages=args.library_max_pages,
            force_fetch=args.force_fetch,
        )
        discovered = merge_discovered(library_discovered, discovered)
        print(
            "加入高中作文文库后："
            + ", ".join(f"{code}={len(rows)}" for code, rows in discovered.items()),
            flush=True,
        )
    if not args.skip_middle_school_backup:
        middle_school_rows = discover_middle_school_ny2_candidates(
            client,
            search_pages=args.middle_school_search_pages,
            category_tail_pages=args.middle_school_category_tail_pages,
            category_page_step=args.middle_school_category_page_step,
            force_fetch=args.force_fetch,
        )
        middle_school_discovered = {code: [] for code in TOPIC_RULES}
        middle_school_discovered["NY2"] = middle_school_rows
        discovered = merge_discovered(discovered, middle_school_discovered)
        print(f"加入初中备选后：NY2={len(discovered['NY2'])}", flush=True)
    pools, audit = build_eligible_pools(
        client=client,
        discovered=discovered,
        manual_reviews=manual_reviews,
        target_per_group=args.target_per_group,
        pool_multiplier=args.pool_multiplier,
        preferred_start=args.preferred_start_year,
        preferred_end=args.preferred_end_year,
        min_han=args.min_han,
        max_han=math.floor(args.max_han * (1 + args.max_han_tolerance)),
        force_fetch=args.force_fetch,
    )
    deduplicated: dict[str, list[CandidateAudit]] = {}
    for code, pool in pools.items():
        kept, rejected = deduplicate_pool(pool)
        deduplicated[code] = kept
        print(f"{code} 去重：{len(pool)} -> {len(kept)}", flush=True)
        for item in rejected:
            if item not in audit:
                audit.append(item)
    deduplicated, cross_rejected = deduplicate_across_groups(deduplicated)
    if cross_rejected:
        print(f"跨组去重：排除{len(cross_rejected)}篇", flush=True)
    available = min((len(pool) for pool in deduplicated.values()), default=0)
    final_count = min(args.target_per_group, available)
    if final_count <= 0:
        raise RuntimeError("没有获得可用的四组共同样本")
    if final_count < args.target_per_group:
        print(
            f"至少一组不足{args.target_per_group}篇，按方案统一缩减为每组{final_count}篇。",
            flush=True,
        )
    selected_by_code = {
        code: select_group(
            pool,
            learner_lengths[TOPIC_RULES[code].learner_code],
            final_count,
            args.preferred_start_year,
            args.preferred_end_year,
        )
        for code, pool in deduplicated.items()
    }
    payload = write_outputs(
        selected_by_code=selected_by_code,
        audit=audit,
        learner_lengths=learner_lengths,
        ori_dir=ori_dir,
        clean_dir=clean_dir,
        output_dir=output_dir,
        preferred_start=args.preferred_start_year,
        preferred_end=args.preferred_end_year,
        nominal_min_han=args.min_han,
        nominal_max_han=args.max_han,
        dry_run=args.dry_run,
    )
    print(
        f"完成：每组{final_count}篇，总计{len(payload['selected'])}篇；"
        f"实时请求{client.live_requests}，缓存命中{client.cache_hits}。",
        flush=True,
    )
    if args.dry_run:
        print("dry-run：未写入语料和审计文件。", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
