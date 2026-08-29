#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import platform
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from linguistic_features import (
    CATEGORY_COHESION,
    CATEGORY_GRAMMAR,
    CATEGORY_HSK,
    CATEGORY_LEXICAL_DENSITY,
    CATEGORY_LEXICAL_DIVERSITY,
    CATEGORY_NARRATIVE,
    CATEGORY_STRUCTURE,
    CONNECTIVE_FEATURES,
    CompiledLexiconEntry,
    FieldSpec,
    RAW_POS_PARENT,
    Token,
    compile_feature_lexicon,
    compute_linguistic_features,
    raw_pos_parent,
    read_feature_lexicon,
)


DEFAULT_WORKBOOK = "作文样本主表.xlsx"
DEFAULT_INPUT_DIR = "clean_text"
DEFAULT_SEG_OUTPUT_DIR = "seg_text"
DEFAULT_STATS_OUTPUT = "outputs/作文词性统计宽表.xlsx"
DEFAULT_HSK_VOCAB = "outputs/新版HSK词汇大纲.csv"
DEFAULT_FEATURE_LEXICON = "resources/语言特征词表.csv"
DEFAULT_MATTR_WINDOW = 50
DEFAULT_LONG_SENTENCE_THRESHOLD = 30
SHEET_NAMES = ("J1", "J2", "Y1", "Y2")
REQUIRED_COLUMNS = ("篇名代码", "篇名", "作文编码", "国籍", "作文题目", "作文分数", "体裁", "作文文件名")
EXPECTED_TOTAL = 620
EXPECTED_HSK_VOCAB_COUNT = 11000
PUNCT_POS = "punctuation mark"
PARENT_POS_TO_RAW_ROOT = {parent: root for root, parent in RAW_POS_PARENT.items()}
HAN_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
HSK_POS_RE = re.compile(r"前缀|后缀|拟声|数量|名|动|形|副|代|介|连|助|数|量|叹")

HSK_LEVELS = ("1", "2", "3", "4", "5", "6", "7-9")
HSK_LEVEL_RANK = {level: index for index, level in enumerate(HSK_LEVELS)}
HSK_LEVEL_GROUPS = {
    "初等": ("1", "2", "3"),
    "中等": ("4", "5", "6"),
    "高等": ("7-9",),
}
HSK_REQUIRED_COLUMNS = ("词语", "主等级", "词性")

PY_NLPIR_TO_HSK_POS = {
    "noun": frozenset({"名"}),
    "pronoun": frozenset({"代"}),
    "verb": frozenset({"动"}),
    "adjective": frozenset({"形"}),
    "adverb": frozenset({"副"}),
    "preposition": frozenset({"介"}),
    "conjunction": frozenset({"连"}),
    "particle": frozenset({"助"}),
    "numeral": frozenset({"数", "数量"}),
    "classifier": frozenset({"量", "数量"}),
    "time word": frozenset({"名"}),
    "noun of locality": frozenset({"名"}),
    "locative word": frozenset({"名"}),
    "suffix": frozenset({"后缀"}),
    "prefix": frozenset({"前缀"}),
    "modal particle": frozenset({"助"}),
    "distinguishing word": frozenset({"形"}),
    "status word": frozenset({"形"}),
    "onomatopoeia": frozenset({"拟声"}),
    "interjection": frozenset({"叹"}),
}

POS_COLUMN_MAP = {
    "noun": "名词数",
    "pronoun": "代词数",
    "verb": "动词数",
    "adjective": "形容词数",
    "adverb": "副词数",
    "preposition": "介词数",
    "conjunction": "连词数",
    "particle": "助词数",
    "numeral": "数词数",
    "classifier": "量词数",
    "time word": "时间词数",
    "noun of locality": "方位词数",
    "locative word": "方位词数",
    "suffix": "后缀数",
    "modal particle": "语气词数",
    "distinguishing word": "区别词数",
    "status word": "状态词数",
    "onomatopoeia": "拟声词数",
    "interjection": "叹词数",
    "multiword expression": "熟语数",
    PUNCT_POS: "标点数",
}

POS_COLUMN_ORDER = [
    "名词数",
    "代词数",
    "动词数",
    "形容词数",
    "副词数",
    "介词数",
    "连词数",
    "助词数",
    "数词数",
    "量词数",
    "时间词数",
    "方位词数",
    "后缀数",
    "语气词数",
    "区别词数",
    "状态词数",
    "拟声词数",
    "叹词数",
    "熟语数",
    "标点数",
]

BASIC_HEADERS = ["篇名代码", "篇名", "作文编码", "国籍", "作文题目", "作文分数", "体裁", "作文文件名"]
TEXT_STAT_HEADERS = ["字数", "纯文本字数", "分词数", "非标点分词数", "去重词数"]
@dataclass(frozen=True)
class EssayRecord:
    sheet_name: str
    code: str
    title_code_name: str
    essay_id: str
    nationality: str
    essay_topic: str
    score: int | str
    genre: str
    filename: str


@dataclass(frozen=True)
class HskVocabularyEntry:
    level: str
    pos_categories: frozenset[str]


@dataclass
class EssayStats:
    record: EssayRecord
    char_count: int
    han_char_count: int
    token_count: int
    non_punct_token_count: int
    unique_word_count: int
    pos_counts: Counter[str]
    hsk_level_counts: Counter[str]
    feature_values: dict[str, int | float]
    feature_fields: tuple[FieldSpec, ...]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "使用 PyNLPIR 对 clean_text 作文分词，并生成词性统计宽表。"
            "请用：arch -x86_64 /usr/bin/python3 segment_hsk_clean_texts.py"
        ),
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help=f"作文主表，默认：{DEFAULT_WORKBOOK}")
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR, help=f"清洗文本目录，默认：{DEFAULT_INPUT_DIR}")
    parser.add_argument("--seg-output-dir", default=DEFAULT_SEG_OUTPUT_DIR, help=f"分词输出目录，默认：{DEFAULT_SEG_OUTPUT_DIR}")
    parser.add_argument("--stats-output", default=DEFAULT_STATS_OUTPUT, help=f"统计宽表输出，默认：{DEFAULT_STATS_OUTPUT}")
    parser.add_argument("--hsk-vocab", default=DEFAULT_HSK_VOCAB, help=f"HSK 词汇表 CSV，默认：{DEFAULT_HSK_VOCAB}")
    parser.add_argument(
        "--feature-lexicon",
        default=DEFAULT_FEATURE_LEXICON,
        help=f"语言特征词表 CSV，默认：{DEFAULT_FEATURE_LEXICON}",
    )
    parser.add_argument(
        "--mattr-window",
        type=int,
        default=DEFAULT_MATTR_WINDOW,
        help=f"MATTR 移动窗口长度，默认：{DEFAULT_MATTR_WINDOW}",
    )
    parser.add_argument(
        "--long-sentence-threshold",
        type=int,
        default=DEFAULT_LONG_SENTENCE_THRESHOLD,
        help=f"长句汉字数阈值（严格大于），默认：{DEFAULT_LONG_SENTENCE_THRESHOLD}",
    )
    parser.add_argument("--limit", type=int, default=None, help="只处理前 N 篇，用于小样本测试")
    parser.add_argument("--dry-run", action="store_true", help="只校验主表和文本文件，不写输出")
    return parser.parse_args()


def require_x86_64() -> None:
    if platform.machine() != "x86_64":
        raise RuntimeError(
            "PyNLPIR 当前只能在 x86_64 Python 下运行。请使用：\n"
            "  arch -x86_64 /usr/bin/python3 segment_hsk_clean_texts.py"
        )


def validate_args(args: argparse.Namespace) -> None:
    if args.limit is not None and args.limit <= 0:
        raise ValueError("--limit 必须大于 0")
    if args.mattr_window <= 0:
        raise ValueError("--mattr-window 必须大于 0")
    if args.long_sentence_threshold <= 0:
        raise ValueError("--long-sentence-threshold 必须大于 0")
    if not str(args.stats_output).lower().endswith(".xlsx"):
        raise ValueError("--stats-output 必须是 .xlsx 文件")


def import_runtime_dependencies():
    try:
        import pynlpir  # type: ignore[import-not-found]
        from openpyxl import Workbook, load_workbook  # type: ignore[import-not-found]
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-not-found]
        from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    except Exception as exc:  # pragma: no cover - user-facing environment guard
        raise RuntimeError(
            "缺少运行依赖。请确认已在 x86_64 Python 环境安装 pynlpir 和 openpyxl：\n"
            "  arch -x86_64 /usr/bin/python3 -m pip install --user pynlpir openpyxl"
        ) from exc
    return pynlpir, Workbook, load_workbook, Alignment, Font, PatternFill, get_column_letter


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()
    return str(value).strip()


def normalize_score(value: Any) -> int | str:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = normalize_cell(value)
    return int(text) if text.isdigit() else text


def read_hsk_vocabulary(path: Path) -> dict[str, tuple[HskVocabularyEntry, ...]]:
    if not path.is_file():
        raise FileNotFoundError(f"找不到 HSK 词汇表：{path}")

    entries_by_word: dict[str, list[HskVocabularyEntry]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        missing_columns = [column for column in HSK_REQUIRED_COLUMNS if column not in headers]
        if missing_columns:
            raise ValueError(f"HSK 词汇表缺少必要列：{missing_columns}；实际表头：{headers}")

        row_count = 0
        for row_number, row in enumerate(reader, start=2):
            row_count += 1
            word = normalize_cell(row.get("词语"))
            level = normalize_cell(row.get("主等级"))
            raw_pos = normalize_cell(row.get("词性"))
            if not word:
                raise ValueError(f"HSK 词汇表第 {row_number} 行缺少词语")
            if level not in HSK_LEVEL_RANK:
                raise ValueError(f"HSK 词汇表第 {row_number} 行主等级无效：{level!r}")

            entries_by_word.setdefault(word, []).append(
                HskVocabularyEntry(
                    level=level,
                    pos_categories=frozenset(HSK_POS_RE.findall(raw_pos.replace(" ", ""))),
                )
            )

    if row_count != EXPECTED_HSK_VOCAB_COUNT:
        raise ValueError(
            f"HSK 词汇表应有 {EXPECTED_HSK_VOCAB_COUNT} 条记录，实际读取 {row_count} 条"
        )
    return {word: tuple(entries) for word, entries in entries_by_word.items()}


def select_hsk_level(
    word: str,
    pos: str | None,
    vocabulary: dict[str, tuple[HskVocabularyEntry, ...]],
) -> str | None:
    entries = vocabulary.get(word)
    if not entries:
        return None

    levels = {entry.level for entry in entries}
    if len(levels) == 1:
        return next(iter(levels))

    target_pos = PY_NLPIR_TO_HSK_POS.get((pos or "").strip(), frozenset())
    matched_levels = {
        entry.level
        for entry in entries
        if target_pos and entry.pos_categories.intersection(target_pos)
    }
    if len(matched_levels) == 1:
        return next(iter(matched_levels))

    return min(levels, key=HSK_LEVEL_RANK.__getitem__)


def read_records(workbook_path: Path, limit: int | None, load_workbook: Any) -> list[EssayRecord]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"找不到作文主表：{workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[EssayRecord] = []
    try:
        missing_sheets = [sheet_name for sheet_name in SHEET_NAMES if sheet_name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError(f"作文主表缺少 sheet：{missing_sheets}")

        for sheet_name in SHEET_NAMES:
            sheet = workbook[sheet_name]
            headers = [normalize_cell(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
            if missing_columns:
                raise ValueError(f"{sheet_name} 缺少必要列：{missing_columns}；实际表头：{headers}")

            index = {column: headers.index(column) for column in REQUIRED_COLUMNS}
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(value is None for value in row):
                    continue

                code = normalize_cell(row[index["篇名代码"]])
                filename = normalize_cell(row[index["作文文件名"]])
                if code != sheet_name:
                    raise ValueError(f"{sheet_name}!A{row_number} 篇名代码应为 {sheet_name}，实际为 {code!r}")
                if not filename:
                    raise ValueError(f"{sheet_name}!{row_number} 缺少作文文件名")
                if filename.endswith(".txt"):
                    filename = filename[:-4]

                records.append(
                    EssayRecord(
                        sheet_name=sheet_name,
                        code=code,
                        title_code_name=normalize_cell(row[index["篇名"]]),
                        essay_id=normalize_cell(row[index["作文编码"]]),
                        nationality=normalize_cell(row[index["国籍"]]),
                        essay_topic=normalize_cell(row[index["作文题目"]]),
                        score=normalize_score(row[index["作文分数"]]),
                        genre=normalize_cell(row[index["体裁"]]),
                        filename=filename,
                    )
                )
    finally:
        workbook.close()

    if len(records) != EXPECTED_TOTAL:
        raise ValueError(f"作文主表应有 {EXPECTED_TOTAL} 条记录，实际读取 {len(records)} 条")
    return records[:limit] if limit is not None else records


def source_path_for(input_dir: Path, record: EssayRecord) -> Path:
    return input_dir / record.code / f"{record.filename}.txt"


def seg_path_for(output_dir: Path, record: EssayRecord) -> Path:
    return output_dir / record.code / f"{record.filename}.txt"


def validate_sources(records: list[EssayRecord], input_dir: Path) -> None:
    missing = [source_path_for(input_dir, record) for record in records if not source_path_for(input_dir, record).exists()]
    if missing:
        preview = "\n".join(str(path) for path in missing[:20])
        raise FileNotFoundError(f"缺少 clean_text 源文件，共 {len(missing)} 个；前几个：\n{preview}")

    duplicate_files = [item for item, count in Counter((record.code, record.filename) for record in records).items() if count > 1]
    if duplicate_files:
        raise ValueError(f"作文文件名重复：{duplicate_files[:20]}")


def pos_column_name(pos: str | None) -> str:
    normalized = (pos or "").strip()
    if normalized in POS_COLUMN_MAP:
        return POS_COLUMN_MAP[normalized]
    safe_pos = re.sub(r"[^0-9A-Za-z]+", "_", normalized).strip("_") or "空"
    return f"其他词性_{safe_pos}数"


def pos_label(pos: str | None) -> str:
    column_name = pos_column_name(pos)
    return column_name[:-1] if column_name.endswith("数") else column_name


def count_chars(text: str) -> tuple[int, int]:
    no_space = re.sub(r"\s+", "", text)
    return len(no_space), len(HAN_RE.findall(no_space))


def segment_lines(
    text: str,
    pynlpir: Any,
    hsk_vocabulary: dict[str, tuple[HskVocabularyEntry, ...]],
) -> tuple[str, Counter[str], Counter[str], int, int, list[Token]]:
    output_lines: list[str] = []
    pos_counts: Counter[str] = Counter()
    hsk_level_counts: Counter[str] = Counter()
    unique_words: set[str] = set()
    feature_tokens: list[Token] = []
    token_count = 0
    paragraph_index = -1

    for line in text.splitlines():
        if not line.strip():
            output_lines.append("")
            continue

        paragraph_index += 1
        segmented = pynlpir.segment(line, pos_names=None)
        parent_fallback: list[tuple[str, str | None]] | None = None
        if any(not normalize_cell(raw_pos) for _, raw_pos in segmented):
            parent_fallback = pynlpir.segment(line)
            raw_words = [normalize_cell(word) for word, _ in segmented]
            parent_words = [normalize_cell(word) for word, _ in parent_fallback]
            if raw_words != parent_words:
                raise ValueError("PyNLPIR原始词性与大类词性两次分词结果不一致，无法安全对齐")
        token_count += len(segmented)
        tokens: list[str] = []
        for token_index, (word, raw_pos) in enumerate(segmented):
            word_text = str(word).replace("\n", " ").strip()
            raw_pos_text = normalize_cell(raw_pos).lower()
            parent_pos = raw_pos_parent(raw_pos_text)
            if not parent_pos and parent_fallback is not None:
                parent_pos = normalize_cell(parent_fallback[token_index][1])
                raw_pos_text = PARENT_POS_TO_RAW_ROOT.get(parent_pos, raw_pos_text)
            label = pos_label(parent_pos)
            pos_counts[pos_column_name(parent_pos)] += 1
            hsk_level: str | None = None
            if parent_pos != PUNCT_POS:
                if word_text:
                    unique_words.add(word_text)
                hsk_level = select_hsk_level(word_text, parent_pos, hsk_vocabulary)
                if hsk_level is not None:
                    hsk_level_counts[hsk_level] += 1
            feature_tokens.append(
                Token(
                    word=word_text,
                    raw_pos=raw_pos_text,
                    parent_pos=parent_pos,
                    paragraph_index=paragraph_index,
                    hsk_level=hsk_level,
                )
            )
            tokens.append(f"{word_text}/{label}")
        output_lines.append(" ".join(tokens))

    return (
        "\n".join(output_lines).rstrip() + "\n",
        pos_counts,
        hsk_level_counts,
        token_count,
        len(unique_words),
        feature_tokens,
    )


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", suffix=".tmp", delete=False) as tmp:
        tmp.write(text)
        tmp_path = Path(tmp.name)
    tmp_path.replace(path)


def segment_record(
    record: EssayRecord,
    input_dir: Path,
    seg_output_dir: Path,
    pynlpir: Any,
    hsk_vocabulary: dict[str, tuple[HskVocabularyEntry, ...]],
    feature_lexicon: tuple[CompiledLexiconEntry, ...],
    mattr_window: int,
    long_sentence_threshold: int,
) -> EssayStats:
    source_path = source_path_for(input_dir, record)
    text = source_path.read_text(encoding="utf-8")
    segmented_text, pos_counts, hsk_level_counts, token_count, unique_word_count, feature_tokens = segment_lines(
        text,
        pynlpir,
        hsk_vocabulary,
    )
    char_count, han_char_count = count_chars(text)
    non_punct_token_count = token_count - pos_counts.get(POS_COLUMN_MAP[PUNCT_POS], 0)
    feature_result = compute_linguistic_features(
        text,
        feature_tokens,
        feature_lexicon,
        mattr_window=mattr_window,
        long_sentence_threshold=long_sentence_threshold,
        hsk_levels=HSK_LEVELS,
        hsk_groups=HSK_LEVEL_GROUPS,
    )
    atomic_write_text(seg_path_for(seg_output_dir, record), segmented_text)
    return EssayStats(
        record=record,
        char_count=char_count,
        han_char_count=han_char_count,
        token_count=token_count,
        non_punct_token_count=non_punct_token_count,
        unique_word_count=unique_word_count,
        pos_counts=pos_counts,
        hsk_level_counts=hsk_level_counts,
        feature_values=feature_result.values,
        feature_fields=feature_result.fields,
    )


CATEGORY_BASIC = "基本信息"
CATEGORY_LENGTH = "基础篇幅"
CATEGORY_POS = "词性"
FEATURE_CATEGORY_ORDER = (
    CATEGORY_LEXICAL_DIVERSITY,
    CATEGORY_LEXICAL_DENSITY,
    CATEGORY_STRUCTURE,
    CATEGORY_GRAMMAR,
    CATEGORY_COHESION,
    CATEGORY_NARRATIVE,
    CATEGORY_HSK,
)


def base_field_specs() -> tuple[FieldSpec, ...]:
    basic_definitions = {
        "篇名代码": "作文所属抽样代码（J1/J2/Y1/Y2）",
        "篇名": "作文任务或篇名",
        "作文编码": "HSK语料库中的作文唯一编码",
        "国籍": "作者国籍",
        "作文题目": "作文题目",
        "作文分数": "作文原始评分",
        "体裁": "作文体裁",
        "作文文件名": "项目内用于关联文本文件的稳定名称",
    }
    fields = [
        FieldSpec(
            name=name,
            category=CATEGORY_BASIC,
            definition=basic_definitions[name],
            formula="来自作文样本主表",
            unit="文本" if name != "作文分数" else "分",
            number_format="@" if name in {"作文编码", "作文文件名"} else "General",
        )
        for name in BASIC_HEADERS
    ]
    length_metadata = {
        "字数": ("去掉空白后的字符总数，包含标点、数字和字母", "去除空白后按Unicode字符计数", "字符"),
        "纯文本字数": ("仅中文汉字字符数量", "匹配Unicode汉字区间后计数", "汉字"),
        "分词数": ("PyNLPIR返回的全部token数量，包含标点", "全部token直接计数", "token"),
        "非标点分词数": ("排除PyNLPIR标点词性后的token数量", "分词数 - 标点数", "token"),
        "去重词数": ("非标点token按词形去重后的数量", "非标点token词形集合大小", "词种"),
    }
    fields.extend(
        FieldSpec(
            name=name,
            category=CATEGORY_LENGTH,
            definition=length_metadata[name][0],
            formula=length_metadata[name][1],
            unit=length_metadata[name][2],
            number_format="0",
        )
        for name in TEXT_STAT_HEADERS
    )
    return tuple(fields)


def pos_rate_name(count_column: str) -> str:
    return f"{count_column[:-1]}每千字" if count_column.endswith("数") else f"{count_column}每千字"


def build_stats_rows(stats: list[EssayStats]) -> tuple[list[FieldSpec], list[list[Any]]]:
    if not stats:
        raise ValueError("没有可写入统计宽表的作文")

    observed_pos_columns = set()
    for item in stats:
        observed_pos_columns.update(item.pos_counts.keys())

    other_pos_columns = sorted(column for column in observed_pos_columns if column not in POS_COLUMN_ORDER)
    pos_columns = list(POS_COLUMN_ORDER) + other_pos_columns

    reference_feature_fields = stats[0].feature_fields
    reference_feature_names = tuple(field.name for field in reference_feature_fields)
    for item in stats[1:]:
        if tuple(field.name for field in item.feature_fields) != reference_feature_names:
            raise ValueError(f"{item.record.filename} 的语言特征字段与首篇不一致")

    feature_fields_by_category: dict[str, list[FieldSpec]] = {category: [] for category in FEATURE_CATEGORY_ORDER}
    for field in reference_feature_fields:
        if field.category not in feature_fields_by_category:
            raise ValueError(f"未知语言特征类别：{field.category}/{field.name}")
        feature_fields_by_category[field.category].append(field)

    fields = list(base_field_specs())
    for category in (CATEGORY_LEXICAL_DIVERSITY, CATEGORY_LEXICAL_DENSITY, CATEGORY_STRUCTURE):
        fields.extend(feature_fields_by_category[category])

    for column in pos_columns:
        label = column[:-1] if column.endswith("数") else column
        fields.append(
            FieldSpec(
                name=column,
                category=CATEGORY_POS,
                definition=f"PyNLPIR大类词性为{label}的token数量",
                formula="按PyNLPIR词性直接计数",
                unit="次",
                number_format="0",
            )
        )
        fields.append(
            FieldSpec(
                name=pos_rate_name(column),
                category=CATEGORY_POS,
                definition=f"{label}token的每千汉字标准化频率",
                formula=f"{column} ÷ 纯文本字数 × 1000",
                unit="次/千汉字",
                number_format="0.00",
            )
        )

    for category in (CATEGORY_GRAMMAR, CATEGORY_COHESION, CATEGORY_NARRATIVE, CATEGORY_HSK):
        fields.extend(feature_fields_by_category[category])

    field_names = [field.name for field in fields]
    duplicate_fields = [name for name, count in Counter(field_names).items() if count > 1]
    if duplicate_fields:
        raise ValueError(f"统计宽表字段名重复：{duplicate_fields}")

    rows: list[list[Any]] = []
    for item in stats:
        record = item.record
        values: dict[str, Any] = {
            "篇名代码": record.code,
            "篇名": record.title_code_name,
            "作文编码": record.essay_id,
            "国籍": record.nationality,
            "作文题目": record.essay_topic,
            "作文分数": record.score,
            "体裁": record.genre,
            "作文文件名": record.filename,
            "字数": item.char_count,
            "纯文本字数": item.han_char_count,
            "分词数": item.token_count,
            "非标点分词数": item.non_punct_token_count,
            "去重词数": item.unique_word_count,
            **item.feature_values,
        }
        for column in pos_columns:
            count = item.pos_counts.get(column, 0)
            values[column] = count
            values[pos_rate_name(column)] = count * 1000 / item.han_char_count if item.han_char_count else 0.0
        rows.append([values[field.name] for field in fields])

    return fields, rows


def save_stats_workbook(
    path: Path,
    stats: list[EssayStats],
    Workbook: Any,
    Alignment: Any,
    Font: Any,
    PatternFill: Any,
    get_column_letter: Any,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields, rows = build_stats_rows(stats)
    headers = [field.name for field in fields]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "词性统计"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    category_colors = {
        CATEGORY_BASIC: "1F4E78",
        CATEGORY_LENGTH: "375623",
        CATEGORY_LEXICAL_DIVERSITY: "8064A2",
        CATEGORY_LEXICAL_DENSITY: "C65911",
        CATEGORY_STRUCTURE: "BF9000",
        CATEGORY_POS: "548235",
        CATEGORY_GRAMMAR: "2F75B5",
        CATEGORY_COHESION: "A64D79",
        CATEGORY_NARRATIVE: "7F6000",
        CATEGORY_HSK: "5B9BD5",
    }
    header_font = Font(color="FFFFFF", bold=True)
    for cell, field in zip(sheet[1], fields):
        cell.fill = PatternFill("solid", fgColor=category_colors[field.category])
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 42

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 10,
        "B": 24,
        "C": 20,
        "D": 14,
        "E": 28,
        "F": 10,
        "G": 10,
        "H": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column_index in range(9, len(headers) + 1):
        header_length = len(str(headers[column_index - 1]))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(header_length + 2, 12), 20)

    for row in sheet.iter_rows(min_row=2):
        for cell, field in zip(row, fields):
            cell.alignment = Alignment(vertical="top")
            cell.number_format = field.number_format

    essay_id_column = headers.index("作文编码") + 1
    for cell in sheet.iter_cols(min_col=essay_id_column, max_col=essay_id_column, min_row=2):
        for item in cell:
            item.number_format = "@"

    dictionary = workbook.create_sheet("字段说明")
    dictionary_headers = ["序号", "字段名", "类别", "定义", "公式", "单位/分母"]
    dictionary.append(dictionary_headers)
    for index, field in enumerate(fields, start=1):
        dictionary.append([index, field.name, field.category, field.definition, field.formula, field.unit])
    for cell in dictionary[1]:
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    dictionary.freeze_panes = "A2"
    dictionary.auto_filter.ref = dictionary.dimensions
    dictionary_widths = (8, 28, 16, 52, 42, 18)
    for index, width in enumerate(dictionary_widths, start=1):
        dictionary.column_dimensions[get_column_letter(index)].width = width
    for row in dictionary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    with NamedTemporaryFile("wb", dir=path.parent, prefix=f".{path.name}.", suffix=".tmp", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    workbook.save(tmp_path)
    tmp_path.replace(path)


def validate_stats(stats: list[EssayStats]) -> None:
    for item in stats:
        summed = sum(item.pos_counts.values())
        punct_count = item.pos_counts.get(POS_COLUMN_MAP[PUNCT_POS], 0)
        if summed != item.token_count:
            raise ValueError(f"{item.record.filename} 分词数不等于词性计数总和：{item.token_count} != {summed}")
        if item.non_punct_token_count != item.token_count - punct_count:
            raise ValueError(f"{item.record.filename} 非标点分词数不等于分词数减标点数")
        if not 0 <= item.unique_word_count <= item.non_punct_token_count:
            raise ValueError(
                f"{item.record.filename} 去重词数超出有效范围："
                f"{item.unique_word_count} > {item.non_punct_token_count}"
            )
        unknown_levels = set(item.hsk_level_counts).difference(HSK_LEVELS)
        if unknown_levels:
            raise ValueError(f"{item.record.filename} 存在无效 HSK 主等级：{sorted(unknown_levels)}")
        hsk_count = sum(item.hsk_level_counts.values())
        if hsk_count > item.non_punct_token_count:
            raise ValueError(
                f"{item.record.filename} HSK 等级词汇次数超过非标点分词数："
                f"{hsk_count} > {item.non_punct_token_count}"
            )

        values = item.feature_values
        if values["HSK词汇次数"] + values["非HSK词汇次数"] != item.non_punct_token_count:
            raise ValueError(f"{item.record.filename} HSK与非HSK次数之和不等于非标点分词数")
        if values["HSK词汇次数"] != hsk_count:
            raise ValueError(f"{item.record.filename} HSK派生总数与原等级统计不一致")
        for level in HSK_LEVELS:
            if values[f"{level}级词汇次数"] != item.hsk_level_counts.get(level, 0):
                raise ValueError(f"{item.record.filename} {level}级词汇次数回归不一致")
            if values[f"{level}级词汇种类数"] > values[f"{level}级词汇次数"]:
                raise ValueError(f"{item.record.filename} {level}级词汇种类数超过次数")
        for group, levels in HSK_LEVEL_GROUPS.items():
            expected = sum(item.hsk_level_counts.get(level, 0) for level in levels)
            if values[f"{group}词汇次数"] != expected:
                raise ValueError(f"{item.record.filename} {group}词汇次数不等于对应数字等级之和")
            if values[f"{group}词汇种类数"] > values[f"{group}词汇次数"]:
                raise ValueError(f"{item.record.filename} {group}词汇种类数超过次数")

        connector_sum = sum(values[f"{feature}数"] for feature in CONNECTIVE_FEATURES)
        if values["连接词总数"] != connector_sum:
            raise ValueError(f"{item.record.filename} 连接词总数不等于八类连接标记之和")
        non_hsk_partition_sum = sum(
            values[f"非HSK{label}数"] for label in ("专名", "数字", "字母串", "其他")
        )
        if values["非HSK词汇次数"] != non_hsk_partition_sum:
            raise ValueError(f"{item.record.filename} 非HSK四类之和不等于非HSK总数")
        if values["词形丰富度TTR"] < 0 or values["词形丰富度TTR"] > 1:
            raise ValueError(f"{item.record.filename} TTR超出0到1范围")
        if values["仅出现一次词数"] > item.unique_word_count:
            raise ValueError(f"{item.record.filename} 仅出现一次词数超过去重词数")


def main() -> int:
    args = parse_args()
    validate_args(args)
    require_x86_64()
    pynlpir, Workbook, load_workbook, Alignment, Font, PatternFill, get_column_letter = import_runtime_dependencies()

    workbook_path = Path(args.workbook)
    input_dir = Path(args.input_dir)
    seg_output_dir = Path(args.seg_output_dir)
    stats_output = Path(args.stats_output)
    hsk_vocab_path = Path(args.hsk_vocab)
    feature_lexicon_path = Path(args.feature_lexicon)

    records = read_records(workbook_path, args.limit, load_workbook)
    validate_sources(records, input_dir)
    hsk_vocabulary = read_hsk_vocabulary(hsk_vocab_path)
    feature_lexicon_specs = read_feature_lexicon(feature_lexicon_path)

    print(f"作文主表：{workbook_path}")
    print(f"清洗文本：{input_dir}")
    print(f"分词输出：{seg_output_dir}")
    print(f"统计宽表：{stats_output}")
    print(f"HSK 词汇表：{hsk_vocab_path}（{sum(len(entries) for entries in hsk_vocabulary.values())} 条）")
    print(f"语言特征词表：{feature_lexicon_path}（{len(feature_lexicon_specs)} 条）")
    print(f"MATTR窗口：{args.mattr_window}；长句阈值：>{args.long_sentence_threshold}字")
    print(f"本次处理：{len(records)} 篇")

    if args.dry_run:
        for record in records[:5]:
            print(f"dry-run: {source_path_for(input_dir, record)} -> {seg_path_for(seg_output_dir, record)}")
        print("dry-run ok")
        return 0

    stats: list[EssayStats] = []
    pynlpir.open()
    try:
        feature_lexicon = compile_feature_lexicon(
            feature_lexicon_specs,
            lambda term: pynlpir.segment(term, pos_names=None),
        )
        for index, record in enumerate(records, start=1):
            item = segment_record(
                record,
                input_dir,
                seg_output_dir,
                pynlpir,
                hsk_vocabulary,
                feature_lexicon,
                args.mattr_window,
                args.long_sentence_threshold,
            )
            stats.append(item)
            hsk_count = sum(item.hsk_level_counts.values())
            print(
                f"[{index}/{len(records)}] {record.code}/{record.filename}: "
                f"{item.token_count} tokens, {item.non_punct_token_count} non-punct, "
                f"{item.unique_word_count} unique, {hsk_count} HSK",
                flush=True,
            )
    finally:
        pynlpir.close()

    validate_stats(stats)
    save_stats_workbook(stats_output, stats, Workbook, Alignment, Font, PatternFill, get_column_letter)
    print(f"完成：分词文件 {len(stats)} 个；统计宽表 {stats_output}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
